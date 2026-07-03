import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from streamlit_extras.metric_cards import style_metric_cards

# ═══════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════F
st.set_page_config(
    page_title="Goal Setting Tool",
    layout="wide",
    page_icon="🎯",
    initial_sidebar_state="collapsed",
)

REQUIRED_COLUMNS = ["Week", "Territory ID", "Territory Name", "Product", "Sales", "Goals"]

TABS = [
    ("Instructions",          "📖"),
    ("Input & Validation",    "📁"),
    ("National Goal Setting", "🎯"),
    ("Final Allocation",      "📊"),
    ("Back Testing",          "🔁"),
]

# ═══════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════
_defaults = {
    "raw_data": None, "data": None, "filtered_data": None,
    "action_type": None, "active_tab": 0, "action_msg": None,
    "nation_goal": None, "nation_goal_submitted": False,
    "ngs_model_tab": None, "eq_pct": 100,
    # Persisted allocation results (used by Final Allocation tab)
    "weighted_result_df": None,
    "weighted_summary":   None,
    "weighted_params":    None,
    "fair_share_result_df": None,
    "equal_result_df":      None,
    "bt_result_df":  None,
    "bt_metrics":    None,
    "bt_sim_summary": None,
    "bt_params_used": None,
    "selected_model_label": None,
    "theme": "dark",
    "uploaded_filename": None, "uploaded_filesize": None,
    "selected_products": None,
    "null_fixed_idx": [],
    "gs_product": None,          # product currently being goal-set
    "gs_completed": [],          # products whose Final Allocation is done
    "bt_intro_seen": False,      # Back Testing intro acknowledged
    "bt_mode": None,             # "manual" | "optimizer"
    "opt_results_df": None,      # grid-search ranking table
    "opt_best_params": None,     # best params from optimizer
    "chosen_params": None,       # params carried into Final Allocation
    "chosen_params_src": None,   # "manual" | "optimizer"
}

for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

theme   = st.session_state.theme
is_dark = theme == "dark"

# ═══════════════════════════════════════════════════════════════════
# ACTIVE TAB
# ═══════════════════════════════════════════════════════════════════
try:
    active = int(st.query_params.get("tab", st.session_state.active_tab))
    active = max(0, min(active, len(TABS) - 1))
except (ValueError, TypeError):
    active = 0
st.session_state.active_tab = active

# ═══════════════════════════════════════════════════════════════════
# THEME PALETTE
# ═══════════════════════════════════════════════════════════════════
if is_dark:
    BG_PAGE      = "#0b1220"
    BG_CARD      = "#111827"
    BG_CARD_SOFT = "#0f172a"
    BG_HERO_FROM = "#0f172a"
    BG_HERO_TO   = "#111827"
    TEXT_PRIMARY = "#f1f5f9"
    TEXT_SECOND  = "#cbd5e1"
    TEXT_MUTED   = "#94a3b8"
    TEXT_CAPTION = "#cbd5e1"  # higher-contrast for descriptions/captions in dark mode
    BORDER       = "#1f2937"
    BORDER_SOFT  = "#1e293b"
    INPUT_BG     = "#0f172a"
    INPUT_BORDER = "#334155"
    EXPANDER_BG  = "#111827"
    INFO_BG      = "#1e3a5f"
    INFO_BORDER  = "#3b82f6"
    SUCCESS_BG   = "#14532d"
    SUCCESS_BORDER = "#22c55e"
    WARNING_BG   = "#7c2d12"
    WARNING_BORDER = "#f97316"
    ERROR_BG     = "#7f1d1d"
    ERROR_BORDER = "#ef4444"
    TABLE_BG       = "#0f172a"
    TABLE_HEADER   = "#1e293b"
    BTN_SOFT_BG    = "#1e293b"
    BTN_SOFT_BORDER= "#334155"
else:
    BG_PAGE      = "#ffffff"
    BG_CARD      = "#ffffff"
    BG_CARD_SOFT = "#f8fafc"
    BG_HERO_FROM = "#f8fafc"
    BG_HERO_TO   = "#eff6ff"
    TEXT_PRIMARY = "#0f172a"
    TEXT_SECOND  = "#475569"
    TEXT_MUTED   = "#64748b"
    TEXT_CAPTION = "#334155"  # higher-contrast for descriptions/captions in light mode
    BORDER       = "#cbd5e1"
    BORDER_SOFT  = "#e2e8f0"
    INPUT_BG     = "#ffffff"
    INPUT_BORDER = "#94a3b8"
    EXPANDER_BG  = "#f8fafc"
    INFO_BG      = "#dbeafe"
    INFO_BORDER  = "#3b82f6"
    SUCCESS_BG   = "#dcfce7"
    SUCCESS_BORDER = "#22c55e"
    WARNING_BG   = "#fef3c7"
    WARNING_BORDER = "#f97316"
    ERROR_BG     = "#fee2e2"
    ERROR_BORDER = "#ef4444"
    TABLE_BG       = "#f8fafc"
    TABLE_HEADER   = "#eff6ff"
    BTN_SOFT_BG    = "#f1f5f9"
    BTN_SOFT_BORDER= "#cbd5e1"

# ── Theme toggle icon ──
_MOON_SVG = (
    "<svg xmlns='http://www.w3.org/2000/svg' width='20' height='20' "
    "viewBox='0 0 24 24' fill='none' stroke='%23ffffff' stroke-width='2' "
    "stroke-linecap='round' stroke-linejoin='round'>"
    "<path d='M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z'/></svg>"
)
_SUN_SVG = (
    "<svg xmlns='http://www.w3.org/2000/svg' width='20' height='20' "
    "viewBox='0 0 24 24' fill='none' stroke='%23ffffff' stroke-width='2' "
    "stroke-linecap='round' stroke-linejoin='round'>"
    "<circle cx='12' cy='12' r='5'/>"
    "<line x1='12' y1='1' x2='12' y2='3'/>"
    "<line x1='12' y1='21' x2='12' y2='23'/>"
    "<line x1='4.22' y1='4.22' x2='5.64' y2='5.64'/>"
    "<line x1='18.36' y1='18.36' x2='19.78' y2='19.78'/>"
    "<line x1='1' y1='12' x2='3' y2='12'/>"
    "<line x1='21' y1='12' x2='23' y2='12'/>"
    "<line x1='4.22' y1='19.78' x2='5.64' y2='18.36'/>"
    "<line x1='18.36' y1='5.64' x2='19.78' y2='4.22'/></svg>"
)
TOGGLE_ICON_URI = "data:image/svg+xml;utf8," + (_MOON_SVG if theme == "light" else _SUN_SVG)

st.markdown(f"""
<style>

/* ── Hide ALL Streamlit chrome ── */
[data-testid="stHeader"]                       {{ display: none !important; }}
[data-testid="stSidebar"]                      {{ display: none !important; }}
[data-testid="stSidebarNav"]                   {{ display: none !important; }}
[data-testid="collapsedControl"]               {{ display: none !important; }}
[data-testid="stToolbar"]                      {{ display: none !important; }}
[data-testid="stDecoration"]                   {{ display: none !important; }}
section[data-testid="stSidebar"]               {{ display: none !important; width: 0 !important; }}
button[kind="header"]                          {{ display: none !important; }}

[data-testid="stAppViewContainer"] > .main     {{ padding-top: 0rem; }}

/* ── Page background ── */
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="stMainBlockContainer"] {{
    background: {BG_PAGE} !important;
    color: {TEXT_PRIMARY} !important;
}}

/* ── Main content area ── */
.block-container {{
    padding-top: 6rem !important;
    padding-left: 4rem !important;
    padding-right: 4rem !important;
    padding-bottom: 3rem !important;
    max-width: 1400px !important;
    margin: 0 auto !important;
    background: {BG_PAGE} !important;
}}

@media (max-width: 1200px) {{
    .block-container {{ padding-left: 2.5rem !important; padding-right: 2.5rem !important; }}
}}
@media (max-width: 768px) {{
    .block-container {{ padding-left: 1.25rem !important; padding-right: 1.25rem !important; }}
}}

/* ── TEXT THEMING ── */
h1, h2, h3, h4, h5, h6,
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4,
[data-testid="stMarkdownContainer"] h5,
[data-testid="stMarkdownContainer"] h6 {{
    color: {TEXT_PRIMARY} !important;
}}

.stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown span, .stMarkdown div,
[data-testid="stMarkdownContainer"], [data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li, [data-testid="stMarkdownContainer"] span,
[data-testid="stMarkdownContainer"] strong, [data-testid="stMarkdownContainer"] em {{
    color: {TEXT_PRIMARY} !important;
}}

.stCaption, [data-testid="stCaptionContainer"], [data-testid="stCaptionContainer"] p {{
    color: {TEXT_CAPTION} !important;
    font-size: 0.88rem !important;
}}

[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] label, .stRadio label, .stCheckbox label,
[data-baseweb="form-control-label"], label {{
    color: {TEXT_PRIMARY} !important;
}}

/* ── INPUT WIDGETS ── */
.stTextInput input, .stNumberInput input, .stDateInput input,
[data-baseweb="input"] input, [data-baseweb="input"],
[data-baseweb="select"] > div {{
    background: {INPUT_BG} !important;
    color: {TEXT_PRIMARY} !important;
    border-color: {INPUT_BORDER} !important;
    caret-color: {TEXT_PRIMARY} !important;
}}

/* Blinking cursor for all text/number inputs */
.stTextInput input,
.stNumberInput input,
.stDateInput input,
[data-baseweb="input"] input {{
    caret-color: {TEXT_PRIMARY} !important;
}}

/* Placeholder text — visible in both light & dark mode */
.stTextInput input::placeholder,
.stNumberInput input::placeholder,
.stDateInput input::placeholder,
[data-baseweb="input"] input::placeholder {{
    color: {TEXT_MUTED} !important;
    opacity: 1 !important;
    -webkit-text-fill-color: {TEXT_MUTED} !important;
}}

[data-baseweb="select"] [class*="ValueContainer"],
[data-baseweb="select"] [class*="SingleValue"],
[data-baseweb="select"] span {{
    color: {TEXT_PRIMARY} !important;
}}

/* ── FILE UPLOADER ── */
[data-testid="stFileUploader"] section,
[data-testid="stFileUploaderDropzone"] {{
    background: {INPUT_BG} !important;
    border: 1px solid {INPUT_BORDER} !important;
    border-radius: 8px !important;
    padding: 1.25rem 1.5rem !important;
    display: flex !important;
    align-items: center !important;
    gap: 1rem !important;
}}

[data-testid="stFileUploaderDropzoneInstructions"],
[data-testid="stFileUploaderDropzoneInstructions"] *,
[data-testid="stFileUploader"] section span,
[data-testid="stFileUploader"] section small {{
    color: {TEXT_PRIMARY} !important;
}}
[data-testid="stFileUploaderDropzoneInstructions"] small,
[data-testid="stFileUploader"] section small {{
    color: {TEXT_MUTED} !important;
    font-size: 12px !important;
}}

[data-testid="stFileUploaderDropzone"] svg,
[data-testid="stFileUploader"] section > svg:first-child {{
    color: {TEXT_MUTED} !important;
    fill: {TEXT_MUTED} !important;
}}

[data-testid="stFileUploaderDropzone"] button {{
    background: transparent !important;
    color: {TEXT_PRIMARY} !important;
    border: 1px solid {INPUT_BORDER} !important;
    font-weight: 500 !important;
    padding: 0.5rem 1.25rem !important;
    border-radius: 6px !important;
    margin-left: auto !important;
}}
[data-testid="stFileUploaderDropzone"] button *,
[data-testid="stFileUploaderDropzone"] button p,
[data-testid="stFileUploaderDropzone"] button span {{
    color: {TEXT_PRIMARY} !important;
}}
[data-testid="stFileUploaderDropzone"] button:hover {{
    background: {BG_CARD_SOFT} !important;
    border-color: {TEXT_MUTED} !important;
}}

[data-testid="stFileUploaderFile"] {{
    background: transparent !important;
    border: none !important;
    padding: 12px 4px !important;
    margin: 0 !important;
    display: flex !important;
    align-items: center !important;
    gap: 10px !important;
}}

[data-testid="stFileUploaderFile"] > svg:first-child,
[data-testid="stFileUploaderFile"] [data-testid="stFileUploaderFileIcon"],
[data-testid="stFileUploaderFile"] > div:first-child > svg {{
    color: {TEXT_MUTED} !important;
    fill: {TEXT_MUTED} !important;
    stroke: {TEXT_MUTED} !important;
    width: 22px !important;
    height: 22px !important;
    flex-shrink: 0 !important;
}}

[data-testid="stFileUploaderFileData"] {{
    display: flex !important;
    align-items: baseline !important;
    gap: 8px !important;
    flex: 1 !important;
    min-width: 0 !important;
}}
[data-testid="stFileUploaderFileData"] *,
[data-testid="stFileUploaderFileName"],
[data-testid="stFileUploaderFileName"] * {{
    color: {TEXT_PRIMARY} !important;
    -webkit-text-fill-color: {TEXT_PRIMARY} !important;
    background: transparent !important;
    opacity: 1 !important;
    visibility: visible !important;
}}
[data-testid="stFileUploaderFileName"] {{
    font-weight: 500 !important;
    font-size: 14px !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}}
[data-testid="stFileUploaderFileData"] small {{
    color: {TEXT_MUTED} !important;
    -webkit-text-fill-color: {TEXT_MUTED} !important;
    font-size: 12px !important;
    flex-shrink: 0 !important;
}}

[data-testid="stFileUploaderFile"] button:nth-of-type(2),
[data-testid="stFileUploader"] button[kind="primary"]:not([data-testid="stFileUploaderDropzone"] button) {{
    display: none !important;
}}

[data-testid="stFileUploaderFile"] button,
[data-testid="stFileUploaderDeleteBtn"] button {{
    width: 28px !important;
    height: 28px !important;
    min-width: 28px !important;
    min-height: 28px !important;
    max-width: 28px !important;
    max-height: 28px !important;
    border-radius: 50% !important;
    background: transparent !important;
    border: none !important;
    padding: 0 !important;
    margin: 0 0 0 auto !important;
    cursor: pointer !important;
    transition: background 0.15s !important;
    font-size: 0 !important;
    color: transparent !important;
    overflow: hidden !important;
    flex-shrink: 0 !important;
    background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='{TEXT_MUTED.replace('#', '%23')}' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><line x1='18' y1='6' x2='6' y2='18'/><line x1='6' y1='6' x2='18' y2='18'/></svg>") !important;
    background-repeat: no-repeat !important;
    background-position: center !important;
    background-size: 14px 14px !important;
}}

[data-testid="stFileUploaderFile"] button > *,
[data-testid="stFileUploaderDeleteBtn"] button > * {{
    display: none !important;
}}

[data-testid="stFileUploaderFile"] button:hover,
[data-testid="stFileUploaderDeleteBtn"] button:hover {{
    background-color: {BG_CARD_SOFT} !important;
    background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='{TEXT_PRIMARY.replace('#', '%23')}' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><line x1='18' y1='6' x2='6' y2='18'/><line x1='6' y1='6' x2='18' y2='18'/></svg>") !important;
}}

.stSlider [data-baseweb="slider"] div {{
    color: {TEXT_PRIMARY} !important;
}}

.stRadio [role="radiogroup"] label {{
    color: {TEXT_PRIMARY} !important;
}}

/* ── CONTAINERS ── */
[data-testid="stExpander"],
.streamlit-expanderHeader,
.streamlit-expanderContent {{
    background: {EXPANDER_BG} !important;
    border-color: {BORDER} !important;
}}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] details,
[data-testid="stExpander"] p,
[data-testid="stExpander"] div,
[data-testid="stExpander"] span {{
    color: {TEXT_PRIMARY} !important;
}}

[data-testid="stAlert"] {{
    background: {INFO_BG} !important;
    border-left: 4px solid {INFO_BORDER} !important;
    color: {TEXT_PRIMARY} !important;
}}
[data-testid="stAlert"] * {{
    color: {TEXT_PRIMARY} !important;
}}

[data-testid="stAlert"][data-baseweb="notification"][kind="success"],
div[data-testid="stAlert"]:has(svg[data-testid="stIconCheck"]) {{
    background: {SUCCESS_BG} !important;
    border-left-color: {SUCCESS_BORDER} !important;
}}

div[data-testid="stAlert"]:has(svg[data-testid="stIconWarn"]) {{
    background: {WARNING_BG} !important;
    border-left-color: {WARNING_BORDER} !important;
}}

div[data-testid="stAlert"]:has(svg[data-testid="stIconError"]) {{
    background: {ERROR_BG} !important;
    border-left-color: {ERROR_BORDER} !important;
}}

hr, [data-testid="stDivider"], [data-testid="stHorizontalBlock"] hr {{
    border-color: {BORDER} !important;
    background: {BORDER} !important;
}}

[data-testid="stProgress"] > div > div {{
    background: {BORDER_SOFT} !important;
}}

/* ── BUTTONS ── */
.stButton button, .stDownloadButton button {{
    background: {BTN_SOFT_BG} !important;
    color: {TEXT_PRIMARY} !important;
    border: 1px solid {BTN_SOFT_BORDER} !important;
    font-weight: 500 !important;
}}
.stButton button p, .stDownloadButton button p,
.stButton button span, .stDownloadButton button span,
.stButton button div, .stDownloadButton button div {{
    color: {TEXT_PRIMARY} !important;
}}
.stButton button:hover, .stDownloadButton button:hover {{
    background: {BG_CARD_SOFT} !important;
    border-color: {INPUT_BORDER} !important;
}}

.stButton button[kind="primary"],
.stButton button[kind="primaryFormSubmit"],
button[data-testid="baseButton-primary"],
button[data-testid="stBaseButton-primary"] {{
    background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
    color: #ffffff !important;
    border: none !important;
    font-weight: 600 !important;
}}
.stButton button[kind="primary"] *,
.stButton button[kind="primary"] p,
.stButton button[kind="primary"] span,
.stButton button[kind="primary"] div,
button[data-testid="baseButton-primary"] *,
button[data-testid="stBaseButton-primary"] * {{
    color: #ffffff !important;
}}
.stButton button[kind="primary"]:hover,
button[data-testid="baseButton-primary"]:hover,
button[data-testid="stBaseButton-primary"]:hover {{
    background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
    color: #ffffff !important;
}}

/* ── NAVBAR ── */
.gst-navbar {{
    position: fixed !important;
    top: 0 !important;
    left: 0 !important;
    right: 0 !important;
    z-index: 999999 !important;
    width: 100%;
    display: grid;
    grid-template-columns: 1fr auto 1fr;
    align-items: center;
    padding: 0 4rem;
    background: #0f172a;
    border-bottom: 1px solid #1e293b;
    height: 76px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.18);
}}

@media (max-width: 1200px) {{
    .gst-navbar {{ padding: 0 2.5rem; }}
}}
@media (max-width: 768px) {{
    .gst-navbar {{ padding: 0 1.25rem; }}
}}

.gst-brand-slot {{ display: flex; justify-content: flex-start; align-items: center; }}

.gst-brand {{
    display: flex;
    align-items: center;
    gap: 12px;
    text-decoration: none !important;
    white-space: nowrap;
}}
.gst-brand:hover, .gst-brand:visited, .gst-brand:focus {{
    text-decoration: none !important;
    color: inherit !important;
}}

.brand-logo {{
    height: 40px;
    width: auto;
    max-width: 180px;
    object-fit: contain;
    display: block;
}}

.brand-icon {{
    width: 38px;
    height: 38px;
    background: linear-gradient(135deg, #2563eb, #1d4ed8);
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 18px;
    box-shadow: 0 4px 12px rgba(37,99,235,0.35);
    flex-shrink: 0;
}}

.brand-text {{ display: flex; flex-direction: column; line-height: 1.2; }}
.brand-title    {{ color: #f1f5f9; font-size: 15px; font-weight: 700; letter-spacing: -0.2px; }}
.brand-subtitle {{ color: #64748b; font-size: 11.5px; font-weight: 500; margin-top: 1px; }}

.gst-tabs {{
    display: flex;
    align-items: center;
    gap: 6px;
    justify-content: center;
}}

.gst-tab {{
    display: inline-flex;
    align-items: center;
    gap: 8px;
    font-size: 14px;
    font-weight: 500;
    padding: 9px 18px;
    border-radius: 9px;
    white-space: nowrap;
    transition: background 0.15s, color 0.15s;
    text-decoration: none !important;
    color: #94a3b8 !important;
    cursor: pointer;
}}
.gst-tab:hover {{
    background: #1e293b;
    color: #e2e8f0 !important;
    text-decoration: none !important;
}}
.gst-tab.active {{
    background: linear-gradient(135deg, #2563eb, #1d4ed8);
    color: #ffffff !important;
    font-weight: 600;
    box-shadow: 0 4px 14px rgba(37,99,235,0.32);
    text-decoration: none !important;
}}
.tab-icon {{ font-size: 14px; opacity: 0.95; }}

.gst-right-slot {{
    display: flex;
    justify-content: flex-end;
    align-items: center;
}}

/* ── METRIC CARDS ── */
[data-testid="stMetric"] {{ color: {TEXT_PRIMARY} !important; }}
[data-testid="stMetric"] label,
[data-testid="stMetricLabel"],
[data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] div {{
    color: {TEXT_SECOND} !important;
}}
[data-testid="stMetricValue"],
[data-testid="stMetricValue"] div,
[data-testid="stMetricValue"] p {{
    color: {TEXT_PRIMARY} !important;
}}
[data-testid="stMetricDelta"],
[data-testid="stMetricDelta"] div,
[data-testid="stMetricDelta"] p,
[data-testid="stMetricDelta"] svg {{
    color: {TEXT_MUTED} !important;
}}

/* ── THEME TOGGLE ── */
#gst-theme-anchor {{ display: none !important; }}

.st-key-theme_toggle_btn {{
    position: fixed !important;
    top: 18px !important;
    right: 4rem !important;
    z-index: 9999999 !important;
    width: 40px !important;
    height: 40px !important;
    margin: 0 !important;
    padding: 0 !important;
}}
@media (max-width: 1200px) {{
    .st-key-theme_toggle_btn {{ right: 2.5rem !important; }}
}}
@media (max-width: 768px) {{
    .st-key-theme_toggle_btn {{ right: 1.25rem !important; }}
}}

.st-key-theme_toggle_btn button {{
    width: 40px !important;
    height: 40px !important;
    min-width: 40px !important;
    min-height: 40px !important;
    border-radius: 50% !important;
    background-color: transparent !important;
    border: 1.5px solid #ffffff !important;
    padding: 0 !important;
    transition: background-color 0.18s, transform 0.18s, border-color 0.18s !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    box-shadow: none !important;
    font-size: 0 !important;
    color: transparent !important;
    overflow: hidden !important;
    background-image: url("{TOGGLE_ICON_URI}") !important;
    background-repeat: no-repeat !important;
    background-position: center !important;
    background-size: 20px 20px !important;
}}

.st-key-theme_toggle_btn button:hover {{
    background-color: rgba(255,255,255,0.12) !important;
    border-color: #ffffff !important;
    transform: scale(1.08) !important;
}}

.st-key-theme_toggle_btn button *,
.st-key-theme_toggle_btn button p,
.st-key-theme_toggle_btn button span,
.st-key-theme_toggle_btn button div {{
    font-size: 0 !important;
    color: transparent !important;
    visibility: hidden !important;
}}

/* ── INSTRUCTIONS HERO + STEPS ── */
.instructions-hero {{
    background: linear-gradient(135deg, {BG_HERO_FROM} 0%, {BG_HERO_TO} 100%);
    border: 1px solid {BORDER};
    border-radius: 16px;
    padding: 2.5rem 2.5rem 2rem;
    margin-bottom: 1.5rem;
    text-align: center;
}}
.instructions-hero h1 {{
    font-size: 32px;
    font-weight: 700;
    color: {TEXT_PRIMARY} !important;
    margin: 0 0 0.5rem;
    letter-spacing: -0.5px;
}}
.instructions-hero p {{
    font-size: 15px;
    color: {TEXT_SECOND} !important;
    margin: 0;
    max-width: 640px;
    margin-left: auto;
    margin-right: auto;
    line-height: 1.6;
}}

.steps-list {{
    background: {BG_CARD};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 1.5rem 2rem;
    margin-bottom: 1.5rem;
}}
.step-item {{
    display: flex;
    gap: 16px;
    padding: 12px 0;
    border-bottom: 1px solid {BORDER_SOFT};
}}
.step-item:last-child {{ border-bottom: none; }}
.step-num {{
    flex-shrink: 0;
    width: 32px;
    height: 32px;
    background: linear-gradient(135deg, #2563eb, #1d4ed8);
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 14px;
    font-weight: 700;
    box-shadow: 0 2px 6px rgba(37,99,235,0.25);
}}
.step-text {{
    font-size: 14.5px;
    color: {TEXT_SECOND} !important;
    line-height: 1.65;
    padding-top: 4px;
}}
.step-text strong {{ color: {TEXT_PRIMARY} !important; }}
.step-text em {{
    color: {TEXT_MUTED} !important;
    font-style: normal;
    font-size: 13.5px;
}}

</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# NAVBAR
# ═══════════════════════════════════════════════════════════════════
import base64
import os

def _load_logo_b64():
    logo_path = os.path.join("assets", "logo.png")
    if not os.path.isfile(logo_path):
        return None
    try:
        with open(logo_path, "rb") as f:
            return "data:image/png;base64," + base64.b64encode(f.read()).decode()
    except Exception:
        return None

LOGO_DATA_URI = _load_logo_b64()


def render_navbar(active_idx: int):
    tabs_html = "".join(
        f'<a href="?tab={i}" target="_top" '
        f'class="gst-tab{" active" if i == active_idx else ""}" '
        f'data-tab="{i}">'
        f'<span class="tab-icon">{icon}</span>{name}</a>'
        for i, (name, icon) in enumerate(TABS)
    )

    if LOGO_DATA_URI:
        brand_inner = f'<img src="{LOGO_DATA_URI}" alt="Company logo" class="brand-logo" />'
    else:
        brand_inner = (
            '<div class="brand-icon">🎯</div>'
            '<div class="brand-text">'
            '<span class="brand-title">Goal Setting Tool</span>'
            '<span class="brand-subtitle">Sales Allocation Platform</span>'
            '</div>'
        )

    st.markdown(f"""
    <nav class="gst-navbar">
        <div class="gst-brand-slot">
            <a class="gst-brand" href="?tab=0" target="_top" data-tab="0">
                {brand_inner}
            </a>
        </div>
        <div class="gst-tabs">{tabs_html}</div>
        <div class="gst-right-slot"></div>
    </nav>

    <script>
    (function() {{
        // Resolve the document that actually holds the navbar.
        // On local Streamlit this is window.parent.document; on Streamlit
        // Cloud the app is nested deeper inside an iframe, so we walk
        // upwards until we find a document with our navbar, falling back
        // to the current document if cross-origin access is blocked.
        function findNav() {{
            const candidates = [];
            try {{ candidates.push(window.top.document); }} catch (e) {{}}
            try {{ candidates.push(window.parent.document); }} catch (e) {{}}
            candidates.push(window.document);

            for (const doc of candidates) {{
                if (!doc) continue;
                const found = doc.querySelector('.gst-navbar');
                if (found) return {{ doc, nav: found }};
            }}
            return null;
        }}

        // Resolve the window we should navigate. Prefer the top-level
        // window so the URL change actually reflects in the address bar;
        // fall back to the current window if blocked.
        function findTargetWindow() {{
            try {{
                void window.top.location.href;   // throws on cross-origin
                return window.top;
            }} catch (e) {{}}
            try {{
                void window.parent.location.href;
                return window.parent;
            }} catch (e) {{}}
            return window;
        }}

        const result = findNav();
        if (!result) return;
        const {{ nav }} = result;
        if (nav.dataset.bound === '1') return;
        nav.dataset.bound = '1';

        nav.addEventListener('click', function(e) {{
            const link = e.target.closest('a[data-tab]');
            if (!link) return;
            e.preventDefault();
            e.stopPropagation();

            const tab = link.dataset.tab;
            const targetWin = findTargetWindow();

            try {{
                const url = new URL(targetWin.location.href);
                url.searchParams.set('tab', tab);
                targetWin.location.href = url.toString();
            }} catch (err) {{
                // Last-resort fallback: let the anchor's target="_top"
                // attribute handle the navigation natively.
                window.location.href = link.getAttribute('href');
            }}
        }}, true);
    }})();
    </script>
    """, unsafe_allow_html=True)

render_navbar(active)

# ═══════════════════════════════════════════════════════════════════
# THEME TOGGLE
# ═══════════════════════════════════════════════════════════════════
toggle_icon = "🌙" if theme == "light" else "☀️"

st.markdown(
    f'<div id="gst-theme-anchor" data-current-theme="{theme}"></div>',
    unsafe_allow_html=True,
)

if st.button(" ", key="theme_toggle_btn",
             help="Switch to dark mode" if theme == "light" else "Switch to light mode"):
    st.session_state.theme = "dark" if theme == "light" else "light"
    st.rerun()

# ═══════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════

def go_to_tab(idx: int):
    st.query_params["tab"] = str(idx)
    st.session_state.active_tab = idx
    st.rerun()

def load_file(f):
    return pd.read_excel(f) if f.name.endswith("xlsx") else pd.read_csv(f)

def validate_schema(df):
    return [c for c in REQUIRED_COLUMNS if c not in df.columns]

def preprocess(df):
    df = df.copy()
    df.columns = df.columns.str.strip()
    df["Week"] = pd.to_datetime(df["Week"], errors="coerce")
    for col in ["Sales", "Goals"]:
        df[col] = (
            df[col].astype(str)
            .str.replace("₹", "", regex=False)
            .str.replace("$", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.strip()
        )
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def apply_date_filter(df, start, end):
    return df[
        (df["Week"] >= pd.to_datetime(start)) &
        (df["Week"] <= pd.to_datetime(end))
    ].copy()
    
def apply_product_filter(df, products):
    """Keep only rows whose Product is in `products`. Empty/None → return all."""
    if not products:
        return df.copy()
    return df[df["Product"].isin(products)].copy()


def product_wise_zscores(df):
    """Absolute z-scores of Sales computed WITHIN each Product group."""
    def _z(s):
        sd = s.std()
        if not sd or np.isnan(sd) or sd == 0:
            return pd.Series(np.zeros(len(s)), index=s.index)
        return np.abs((s - s.mean()) / sd)
    if df.empty:
        return pd.Series(dtype=float)
    return df.groupby("Product")["Sales"].transform(_z)


def fill_missing_product_wise(full_df, products, method):
    """
    Fill missing Sales/Goals for the given products, PRODUCT-WISE.
    Mean/Median/Mode use each product's own values only. Previous/Next
    use week-ordered ffill/bfill within each product.
    """
    df   = full_df.copy()
    mask = df["Product"].isin(products)
    cols = ["Sales", "Goals"]

    if method == "Fill with 0":
        df.loc[mask, cols] = df.loc[mask, cols].fillna(0)
        return df

    for col in cols:
        sub = df.loc[mask]
        if method == "Fill with Mean":
            filled = sub.groupby("Product")[col].transform(lambda s: s.fillna(s.mean()))
        elif method == "Fill with Median":
            filled = sub.groupby("Product")[col].transform(lambda s: s.fillna(s.median()))
        elif method == "Fill with Mode":
            filled = sub.groupby("Product")[col].transform(
                lambda s: s.fillna(s.mode().iloc[0]) if not s.mode().empty else s)
        elif method == "Fill using Previous Value":
            filled = sub.sort_values("Week").groupby("Product")[col].ffill()
        elif method == "Fill using Next Value":
            filled = sub.sort_values("Week").groupby("Product")[col].bfill()
        else:
            filled = df.loc[mask, col]
        df.loc[filled.index, col] = filled
    return df


def cap_outliers_product_wise(full_df, products):
    """Cap Sales at each product's own +3σ, only for the selected products."""
    df = full_df.copy()
    n_capped = 0
    for prod in products:
        m  = df["Product"] == prod
        s  = df.loc[m, "Sales"]
        sd = s.std()
        if not sd or np.isnan(sd) or sd == 0:
            continue
        cap = s.mean() + 3 * sd
        n_capped += int((s > cap).sum())
        df.loc[m, "Sales"] = s.clip(upper=cap)
    return df, n_capped

def fmt(n, prefix="$"):
    return "—" if pd.isna(n) else f"{prefix}{n:,.0f}"

def pct_fmt(n):
    return "—" if pd.isna(n) else f"{n:+.1%}"

def to_excel(df, sheet_name="Sheet1", percent_cols=None, currency_cols=None,
             title=None, freeze_header=True):
    """
    Export a DataFrame to a professionally formatted Excel file.

    Styling applied:
      • Bold white header row with dark navy fill
      • Alternating row colors (zebra striping) for readability
      • Thin grey borders on every cell
      • Numeric columns formatted with comma separators + 2 decimal places
      • `percent_cols` formatted as `0.00%` with right-alignment
      • `currency_cols` formatted with ₹ symbol + 2 decimal places
      • Auto-sized column widths based on content
      • Frozen header row
      • Bold + accent fill for any row whose first cell contains "TOTAL"
      • Optional title row at the top (merged across all columns)
    """
    output = BytesIO()
    percent_cols  = percent_cols  or []
    currency_cols = currency_cols or []

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Offset rows down by 2 if a title is supplied (title row + blank row)
        start_row = 3 if title else 1
        df.to_excel(writer, index=False, sheet_name=sheet_name,
                    startrow=start_row - 1)
        ws = writer.sheets[sheet_name]

        # ── Style constants ───────────────────────────────────────
        HEADER_FILL  = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        TITLE_FONT   = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        TITLE_FILL   = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        ROW_FILL_A   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ROW_FILL_B   = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        TOTAL_FILL   = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        TOTAL_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        BODY_FONT    = Font(name="Calibri", size=10, color="0F172A")

        thin   = Side(border_style="thin", color="CBD5E1")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT   = Alignment(horizontal="left",   vertical="center")
        RIGHT  = Alignment(horizontal="right",  vertical="center")

        n_cols = ws.max_column
        n_rows = ws.max_row
        header_row = start_row

        # ── Title row (optional) ──────────────────────────────────
        if title:
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=n_cols)
            tc = ws.cell(row=1, column=1)
            tc.value     = title
            tc.font      = TITLE_FONT
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.fill      = TITLE_FILL
            ws.row_dimensions[1].height = 28

        # ── Header row styling ───────────────────────────────────
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
        ws.row_dimensions[header_row].height = 32

        # Build header → index lookup
        headers = {ws.cell(row=header_row, column=i).value: i
                   for i in range(1, n_cols + 1)}

        # ── Body styling ──────────────────────────────────────────
        first_body_row = header_row + 1
        for r_idx in range(first_body_row, n_rows + 1):
            # Detect totals row (first column says "TOTAL")
            first_cell_val = ws.cell(row=r_idx, column=1).value
            is_total = isinstance(first_cell_val, str) and first_cell_val.strip().upper() == "TOTAL"

            stripe_fill = ROW_FILL_A if (r_idx - first_body_row) % 2 == 0 else ROW_FILL_B
            row_fill    = TOTAL_FILL if is_total else stripe_fill
            row_font    = TOTAL_FONT if is_total else BODY_FONT

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font   = row_font
                cell.fill   = row_fill
                cell.border = BORDER

                if isinstance(cell.value, (int, float)):
                    cell.alignment = RIGHT
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = LEFT if col_idx <= 2 else RIGHT

        # ── Apply percent / currency formats ──────────────────────
        for col_name in percent_cols:
            if col_name in headers:
                idx = headers[col_name]
                for r in range(first_body_row, n_rows + 1):
                    c = ws.cell(row=r, column=idx)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '0.00%'
                        c.alignment = RIGHT

        for col_name in currency_cols:
            if col_name in headers:
                idx = headers[col_name]
                for r in range(first_body_row, n_rows + 1):
                    c = ws.cell(row=r, column=idx)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '"$"#,##0.00'
                        c.alignment = RIGHT

        # ── Auto-size column widths ───────────────────────────────
        for col_idx in range(1, n_cols + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for r in range(header_row, n_rows + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(max_len + 4, 12), 36)

        # ── Freeze pane below header ──────────────────────────────
        if freeze_header:
            ws.freeze_panes = ws.cell(row=first_body_row, column=1)

        # ── Disable Excel gridlines for a cleaner look ────────────
        ws.sheet_view.showGridLines = False

    output.seek(0)
    return output

# ═══════════════════════════════════════════════════════════════════
# CHARTS
# ═══════════════════════════════════════════════════════════════════

def plot_trend(df, level):
    freq_map = {"Weekly": "W", "Monthly": "ME", "Quarterly": "QE"}
    per_map  = {"Weekly": "W", "Monthly": "M",  "Quarterly": "Q"}
    d = df.copy()
    d["Week"] = pd.to_datetime(d["Week"])
    data_max = d["Week"].max()
    agg = (
        d.sort_values("Week")
         .set_index("Week")[["Sales", "Goals"]]
         .resample(freq_map[level])
         .sum()
         .reset_index()
    )
    # Drop any bin whose period is AFTER the last real data point
    # (kills spurious future labels like "Jan 2025" when data ends Dec 2024).
    agg = agg[agg["Week"].dt.to_period(per_map[level]) <= data_max.to_period(per_map[level])]

    axis_color   = "#f1f5f9" if is_dark else "#0f172a"
    grid_color   = "#1f2937" if is_dark else "#e2e8f0"
    tooltip_bg   = "#1e293b" if is_dark else "#ffffff"
    tooltip_fg   = "#f1f5f9" if is_dark else "#0f172a"
    tooltip_brd  = "#334155" if is_dark else "#cbd5e1"

    x_title_map = {"Weekly": "Week", "Monthly": "Month", "Quarterly": "Quarter"}
    x_title = x_title_map.get(level, "Period")

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=agg["Week"], y=agg["Sales"],
        mode="lines+markers", name="Sales",
        line=dict(color="#2563eb", width=3), marker=dict(size=6),
        hovertemplate="<b>%{x|%b %Y}</b><br>Sales: %{y:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=agg["Week"], y=agg["Goals"],
        mode="lines+markers", name="Goals",
        line=dict(color="#f97316", width=3, dash="dash"), marker=dict(size=6),
        hovertemplate="Goals: %{y:,.0f}<extra></extra>",
    ))
    x_range = [agg["Week"].min(), agg["Week"].max()] if not agg.empty else None
    fig.update_layout(
        template="plotly_dark" if is_dark else "plotly_white",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        height=420, hovermode="x unified",
        margin=dict(l=60, r=20, t=30, b=60),
        legend=dict(orientation="h", y=1.05, x=1, xanchor="right",
                    font=dict(color=axis_color)),
        xaxis=dict(
            showgrid=False, color=axis_color,
            range=x_range,                       # ← pins axis to real data
            title=dict(text=x_title,
                       font=dict(color=axis_color, size=14, family="Calibri"),
                       standoff=12),
            tickfont=dict(color=axis_color, size=11),
            linecolor=grid_color,
        ),
        yaxis=dict(
            showgrid=True, tickformat=",.0f", color=axis_color,
            title=dict(text="Sales / Goals ($)",
                       font=dict(color=axis_color, size=14, family="Calibri"),
                       standoff=12),
            tickfont=dict(color=axis_color, size=11),
            gridcolor=grid_color, linecolor=grid_color,
        ),
        font=dict(color=axis_color),
        hoverlabel=dict(bgcolor=tooltip_bg, bordercolor=tooltip_brd,
                        font=dict(color=tooltip_fg, size=12)),
    )
    return fig

def plot_bars(terr, top_n=10):
    t = terr.head(top_n)

    axis_color   = "#f1f5f9" if is_dark else "#0f172a"
    grid_color   = "#1f2937" if is_dark else "#e2e8f0"
    tooltip_bg   = "#1e293b" if is_dark else "#ffffff"
    tooltip_fg   = "#f1f5f9" if is_dark else "#0f172a"
    tooltip_brd  = "#334155" if is_dark else "#cbd5e1"

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=t["Territory Name"], y=t["Sales"], name="Sales",
        marker_color="#3b82f6",
        hovertemplate="%{x}<br>Sales: %{y:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        x=t["Territory Name"], y=t["Goals"], name="Goals",
        marker_color="#f97316", opacity=0.7,
        hovertemplate="%{x}<br>Goals: %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(
        template="plotly_dark" if is_dark else "plotly_white",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        barmode="group", height=300,
        margin=dict(l=10, r=10, t=10, b=80),
        legend=dict(orientation="h", y=1.08, x=1, xanchor="right",
                    font=dict(color=axis_color)),
        xaxis=dict(tickangle=-30, color=axis_color,
                   tickfont=dict(color=axis_color),
                   linecolor=grid_color),
        yaxis=dict(showgrid=True, tickformat=",.0f", color=axis_color,
                   tickfont=dict(color=axis_color),
                   gridcolor=grid_color, linecolor=grid_color),
        font=dict(color=axis_color),
        hoverlabel=dict(
            bgcolor=tooltip_bg,
            bordercolor=tooltip_brd,
            font=dict(color=tooltip_fg, size=12),
        ),
    )
    return fig

# ═══════════════════════════════════════════════════════════════════
# ALLOCATION MODELS
# ═══════════════════════════════════════════════════════════════════

def compute_fair_share(df, nation_goal, mr3_window=3):
    df = df.copy()
    df["Month"] = df["Week"].dt.to_period("M")
    all_months = sorted(df["Month"].unique())
    mr3_months = all_months[-mr3_window:] if len(all_months) >= mr3_window else all_months
    mr3 = (
        df[df["Month"].isin(mr3_months)]
        .groupby(["Territory ID", "Territory Name"])["Sales"]
        .sum().reset_index()
        .rename(columns={"Sales": "MR3"})
    )
    total = mr3["MR3"].sum()
    if total == 0:
        mr3["Share_%"]       = 0
        mr3["Goal"]          = 0
        mr3["Goal_Growth_%"] = np.nan
    else:
        mr3["Share_%"]       = mr3["MR3"] / total
        mr3["Goal"]          = mr3["Share_%"] * nation_goal
        mr3["Goal_Growth_%"] = (mr3["Goal"] - mr3["MR3"]) / mr3["MR3"]
    nation_growth = (nation_goal / total - 1) if total > 0 else np.nan
    return mr3.sort_values("MR3", ascending=False), total, nation_growth, mr3_months


def compute_equal_allocation(df, nation_goal, equal_pct):
    territories = df[["Territory ID", "Territory Name"]].drop_duplicates().copy()
    n = len(territories)
    if n == 0:
        return pd.DataFrame()
    territories["Final_Goal"] = (nation_goal * equal_pct / 100) / n
    # Every territory gets an equal share, so % share = 1/n for all rows
    territories["Share_%"]    = 1.0 / n if n > 0 else 0
    return territories


def compute_weighted_model(df, nation_goal, params):
    """
    Compute territory goals using a weighted blend of recent + older performance,
    with optional volume smoothing, growth cap/floor, and two redistribution passes.

    Parameters:
        hist_scale      Historical Sales scaling factor (default 1.0)
        flat_goal_pct   Equal-share portion of nation goal (default 0.0)
        mr_weight       Recent performance weight (default 0.70)
        p_weight        Older performance weight (auto = 1 − mr_weight)
        vol_balance     Volume smoothing moderator (default 0.015)
        growth_cap      Maximum allowed growth % per territory (default 0.06)
        growth_floor    Minimum allowed growth % per territory (default 0.0)
        mr3_window      # of recent months for MR baseline (default 3)
        p3_window       # of prior months for P  baseline (default 3)

    Computation stages:
        Flat_Goal        = flat_pool / n_territories
        MR3, P3          = recent / older period sales × hist_scale
        Initial_Goal     = (MR3 share × pool × MR_wt) + (P3 share × pool × P_wt)
        Vol_Bal_Adj      = (avg − initial) × vol_balance       (smoothing)
        Goal_After_VolAdj= Initial + smoothing
        Capped_Floored   = clamp by cap/floor + per-territory adjustment
        Redistributed    = re-clamp + per-territory redistribution
        Final_Goal       = Redistributed + Flat_Goal

    Returns: (result_df, summary_dict)
    """
    df = df.copy()
    df["Month"] = df["Week"].dt.to_period("M")
    months_sorted = sorted(df["Month"].unique())

    mr_win = params["mr3_window"]
    p_win  = params["p3_window"]

    if len(months_sorted) < mr_win + p_win:
        mr_months = months_sorted[-mr_win:]
        p_months  = months_sorted[:-mr_win][-p_win:] if len(months_sorted) > mr_win else []
    else:
        mr_months = months_sorted[-mr_win:]
        p_months  = months_sorted[-(mr_win + p_win):-mr_win]

    # ── Per-territory MR3 & P3 ──
    base = (df.groupby(["Territory ID", "Territory Name"])
              .size().reset_index().drop(columns=0))
    mr3 = (df[df["Month"].isin(mr_months)]
           .groupby(["Territory ID", "Territory Name"])["Sales"]
           .sum().reset_index().rename(columns={"Sales": "MR3_raw"}))
    p3  = (df[df["Month"].isin(p_months)]
           .groupby(["Territory ID", "Territory Name"])["Sales"]
           .sum().reset_index().rename(columns={"Sales": "P3_raw"}))

    out = (base.merge(mr3, on=["Territory ID", "Territory Name"], how="left")
                .merge(p3,  on=["Territory ID", "Territory Name"], how="left"))
    out["MR3_raw"] = out["MR3_raw"].fillna(0)
    out["P3_raw"]  = out["P3_raw"].fillna(0)

    # Col I, J — scaled by hist_scale (C18)
    out["MR3"] = out["MR3_raw"] * params["hist_scale"]
    out["P3"]  = out["P3_raw"]  * params["hist_scale"]

    # Col K — Current Growth
    out["Current_Growth"] = np.where(out["P3"] > 0,
                                     (out["MR3"] - out["P3"]) / out["P3"], np.nan)

    n_terr     = len(out)
    flat_pool  = nation_goal * params["flat_goal_pct"]
    weighted_pool = nation_goal - flat_pool

    # Col H — Flat Goal per territory
    out["Flat_Goal"] = flat_pool / n_terr if n_terr > 0 else 0

    # Col L — Initial weighted goal [B]
    total_mr3 = out["MR3"].sum()
    total_p3  = out["P3"].sum()
    if total_mr3 > 0 and total_p3 > 0:
        out["Initial_Goal"] = (
            (out["MR3"] / total_mr3) * weighted_pool * params["mr_weight"] +
            (out["P3"]  / total_p3)  * weighted_pool * params["p_weight"]
        )
    elif total_mr3 > 0:
        out["Initial_Goal"] = (out["MR3"] / total_mr3) * weighted_pool
    else:
        out["Initial_Goal"] = 0

    # Col M
    out["Initial_Growth"] = np.where(out["MR3"] > 0,
                                     (out["Initial_Goal"] - out["MR3"]) / out["MR3"], np.nan)

    # Col N — Volume Balance Adjustment
    avg_initial = out["Initial_Goal"].mean()
    out["Vol_Bal_Adj"] = (avg_initial - out["Initial_Goal"]) * params["vol_balance"]

    # Col O
    out["Goal_After_VolAdj"] = out["Initial_Goal"] + out["Vol_Bal_Adj"]

    # Col P
    out["VolAdj_Growth"] = np.where(out["MR3"] > 0,
                                    (out["Goal_After_VolAdj"] - out["MR3"]) / out["MR3"], np.nan)

    cap   = params["growth_cap"]
    floor = params["growth_floor"]

    # ── Summary R7/R8/R9 (first cap/floor pass) ──
    p_growth = out["VolAdj_Growth"].fillna(0)
    capped_mask  = p_growth >= cap
    floored_mask = p_growth <= floor

    capped_sum_at_cap     = (out.loc[capped_mask,  "MR3"] * (1 + cap)).sum()
    floored_sum_at_floor  = (out.loc[floored_mask, "MR3"] * (1 + floor)).sum()
    middle_sum            = out.loc[~(capped_mask | floored_mask), "Goal_After_VolAdj"].sum()

    R7_adjusted_sales = weighted_pool - (capped_sum_at_cap + floored_sum_at_floor + middle_sum)
    R8_terr_in_band   = int((~capped_mask & ~floored_mask).sum())
    R9_adj_per_terr   = R7_adjusted_sales / R8_terr_in_band if R8_terr_in_band > 0 else 0

    # Col Q — Capped & Floored Goal
    def q_calc(row):
        pg = row["VolAdj_Growth"] if pd.notna(row["VolAdj_Growth"]) else 0
        if pg >= cap or pg <= floor:
            clamped = min(max(pg, floor), cap)
            return row["MR3"] * (1 + clamped)
        else:
            clamped = min(max(pg, floor), cap)
            return row["MR3"] * (1 + clamped) + R9_adj_per_terr
    out["Capped_Floored_Goal"] = out.apply(q_calc, axis=1)

    # Col R
    out["Capped_Growth"] = np.where(out["MR3"] > 0,
                                    (out["Capped_Floored_Goal"] - out["MR3"]) / out["MR3"], np.nan)

    # ── Summary T7/T8/T9 (second pass redistribution) ──
    r_growth = out["Capped_Growth"].fillna(0)
    over_cap_mask  = r_growth >= cap
    under_floor_mask = r_growth <= floor

    cap_capped_at_cap = (out.loc[over_cap_mask, "MR3"] * (1 + cap)).sum()
    cap_capped_actual = out.loc[over_cap_mask, "Capped_Floored_Goal"].sum()
    T7_redistribute   = cap_capped_actual - cap_capped_at_cap

    T8_terr_receiving = int((~(over_cap_mask | under_floor_mask)).sum())
    T9_redist_per_terr = T7_redistribute / T8_terr_receiving if T8_terr_receiving > 0 else 0

    # Col S — Redistributed Goal [D]
    def s_calc(row):
        rg = row["Capped_Growth"] if pd.notna(row["Capped_Growth"]) else 0
        if rg >= cap:
            return row["MR3"] * (1 + cap)
        elif rg <= floor:
            return row["MR3"] * (1 + floor)
        else:
            return row["Capped_Floored_Goal"] + T9_redist_per_terr
    out["Redistributed_Goal"] = out.apply(s_calc, axis=1)

    # ── PRO-RATA FALLBACK ──
    # When EVERY territory is at the cap or at the floor, the standard
    # redistribution mechanism has nobody "in-band" to spread the leftover to,
    # and the difference between Nation Goal and Allocated Total would silently
    # be dropped. To preserve the invariant (Allocated Total ≈ Nation Goal),
    # we spread any remaining gap proportionally across all territories based
    # on their current Redistributed_Goal share. This may push some territories
    # past the user-set cap/floor - the UI shows a warning when this happens.
    fallback_triggered = False
    effective_cap      = cap
    effective_floor    = floor

    current_total = out["Redistributed_Goal"].sum()
    gap = weighted_pool - current_total

    if abs(gap) > 1 and current_total > 0:
        fallback_triggered = True
        # Pro-rata uplift / cutback based on each territory's share
        out["Redistributed_Goal"] = out["Redistributed_Goal"] * (weighted_pool / current_total)

        # Recompute implied growth so we can report the effective cap/floor
        recomputed_growth = np.where(out["MR3"] > 0,
                                     (out["Redistributed_Goal"] - out["MR3"]) / out["MR3"],
                                     np.nan)
        effective_cap   = float(np.nanmax(recomputed_growth)) if len(recomputed_growth) else cap
        effective_floor = float(np.nanmin(recomputed_growth)) if len(recomputed_growth) else floor

    # Col T (recomputed after possible fallback)
    out["Redist_Growth"] = np.where(out["MR3"] > 0,
                                    (out["Redistributed_Goal"] - out["MR3"]) / out["MR3"], np.nan)

    # Col U — Final Goal [E] = S + H
    out["Final_Goal"] = out["Redistributed_Goal"] + out["Flat_Goal"]

    # Col V
    out["Final_Growth"] = np.where(out["MR3"] > 0,
                                   (out["Final_Goal"] - out["MR3"]) / out["MR3"], np.nan)

    out = out.sort_values("MR3", ascending=False).reset_index(drop=True)

    summary = {
        "n_territories":      n_terr,
        "mr_months":          [str(m) for m in mr_months],
        "p_months":           [str(m) for m in p_months],
        "total_mr3":          total_mr3,
        "total_p3":           total_p3,
        "weighted_pool":      weighted_pool,
        "flat_pool":          flat_pool,
        "avg_initial":        avg_initial,
        "R7_adjusted_sales":  R7_adjusted_sales,
        "R8_terr_in_band":    R8_terr_in_band,
        "R9_adj_per_terr":    R9_adj_per_terr,
        "T7_redistribute":    T7_redistribute,
        "T8_terr_receiving":  T8_terr_receiving,
        "T9_redist_per_terr": T9_redist_per_terr,
        "final_total_goal":   out["Final_Goal"].sum(),
        "final_minus_nation": out["Final_Goal"].sum() - nation_goal,
        "fallback_triggered": fallback_triggered,
        "effective_cap":      effective_cap,
        "effective_floor":    effective_floor,
        "user_cap":           cap,
        "user_floor":         floor,
    }
    return out, summary

def compute_back_test(df, test_quarter, prior_quarter, params):
    """
    Inputs:
        df             : raw data with Week/Sales/Goals/Territory
        test_quarter   : pd.Period('YYYY Qq') — the RECENT quarter being back-tested
                         (acts as the "MR3" in the workbook + holds Actual Sales)
        prior_quarter  : pd.Period('YYYY Qq') — the PRIOR quarter (acts as "P3")
        params         : Weighted Model params (same keys as compute_weighted_model)

    Workflow:
        TRAIN data  = all rows on/before the *end* of prior_quarter
                      (so the model has prior_quarter as its "P3" and the
                      quarter just before prior_quarter as its history).
        TEST window = test_quarter (used for Actual Sales + Original Goals).
        Simulated   = compute_weighted_model on TRAIN with National Goal =
                      total actual sales of test_quarter (matches workbook C17/C21).

    Returns: (per_territory_df, metrics_dict, sim_summary_dict)
    """
    df = df.copy()
    df["Week"]    = pd.to_datetime(df["Week"])
    df["Quarter"] = df["Week"].dt.to_period("Q")

    # ── Slice windows ────────────────────────────────────────────
    test_df  = df[df["Quarter"] == test_quarter].copy()
    train_df = df[df["Week"] <= prior_quarter.end_time].copy()
    if test_df.empty or train_df.empty:
        return pd.DataFrame(), {}, {}

    # ── Aggregate the TEST quarter: sales + original goals ───────
    actual = (test_df.groupby(["Territory ID", "Territory Name"])
              .agg(Actual_Sales=("Sales", "sum"),
                   Original_Goal=("Goals", "sum"))
              .reset_index())

    # ── Run the Weighted Model on the TRAIN data ─────────────────
    sim_nation_goal = float(actual["Actual_Sales"].sum())
    sim_result, sim_summary = compute_weighted_model(train_df, sim_nation_goal, params)
    sim = sim_result[["Territory ID", "Territory Name", "Final_Goal", "Final_Growth"]].rename(
        columns={"Final_Goal": "Simulated_Goal", "Final_Growth": "Simulated_Growth"})

    out = actual.merge(sim, on=["Territory ID", "Territory Name"], how="left")
    out[["Simulated_Goal", "Simulated_Growth"]] = out[["Simulated_Goal", "Simulated_Growth"]].fillna(0)

    # ── Per-territory metrics (mirrors workbook cols X, Y, Z, AB, AC, AD, AE, AF, AG, AH) ──
    out["Original_Attainment"]  = np.where(out["Original_Goal"]  > 0,
                                           out["Actual_Sales"] / out["Original_Goal"],  np.nan)
    out["Simulated_Attainment"] = np.where(out["Simulated_Goal"] > 0,
                                           out["Actual_Sales"] / out["Simulated_Goal"], np.nan)
    # Earnings (workbook cols Y/AC) = attainment (capped at 200% for sanity)
    out["Original_Earnings"]    = out["Original_Attainment"].clip(upper=2).fillna(0)
    out["Simulated_Earnings"]   = out["Simulated_Attainment"].clip(upper=2).fillna(0)
    # Payout (workbook cols Z/AD) = earnings × ₹10,000 per territory
    out["Original_Payout"]      = out["Original_Earnings"]  * 10000
    out["Simulated_Payout"]     = out["Simulated_Earnings"] * 10000
    # Absolute Percentage Error (workbook cols AG/AH)
    out["Original_APE"]  = np.where(out["Actual_Sales"] > 0,
                                    (out["Original_Goal"]  - out["Actual_Sales"]).abs() / out["Actual_Sales"], np.nan)
    out["Simulated_APE"] = np.where(out["Actual_Sales"] > 0,
                                    (out["Simulated_Goal"] - out["Actual_Sales"]).abs() / out["Actual_Sales"], np.nan)
    # Squared Error for RMSE (workbook cols AE/AF)
    out["Original_SE"]   = (out["Original_Goal"]  - out["Actual_Sales"]) ** 2
    out["Simulated_SE"]  = (out["Simulated_Goal"] - out["Actual_Sales"]) ** 2

    def _r2(goal_col):
        v = out[[goal_col, "Actual_Sales"]].dropna()
        if len(v) < 2 or v[goal_col].std() == 0 or v["Actual_Sales"].std() == 0:
            return np.nan
        return float(np.corrcoef(v[goal_col], v["Actual_Sales"])[0, 1] ** 2)

    n = len(out)
    metrics = {
        "n_territories":     n,
        "sim_nation_goal":   sim_nation_goal,
        "total_actual":      float(out["Actual_Sales"].sum()),
        "total_original":    float(out["Original_Goal"].sum()),
        "total_simulated":   float(out["Simulated_Goal"].sum()),
        "mape_original":     float(out["Original_APE"].mean())  if n else np.nan,
        "mape_simulated":    float(out["Simulated_APE"].mean()) if n else np.nan,
        "rmse_original":     float(np.sqrt(out["Original_SE"].mean()))  if n else np.nan,
        "rmse_simulated":    float(np.sqrt(out["Simulated_SE"].mean())) if n else np.nan,
        "r2_original":       _r2("Original_Goal"),
        "r2_simulated":      _r2("Simulated_Goal"),
        "payout_original":   float(out["Original_Payout"].sum()),
        "payout_simulated":  float(out["Simulated_Payout"].sum()),
        "budget_payout":     n * 10000,
        "engagement_original":  float((out["Original_Attainment"]  > 0).mean()) if n else np.nan,
        "engagement_simulated": float((out["Simulated_Attainment"] > 0).mean()) if n else np.nan,
        "test_quarter":  str(test_quarter),
        "prior_quarter": str(prior_quarter),
    }
    metrics["winner"] = (
        "Simulated (Weighted Model)"
        if (metrics["mape_simulated"] or 9e9) < (metrics["mape_original"] or 9e9)
        else "Original Goals"
    )
    return out.sort_values("Actual_Sales", ascending=False).reset_index(drop=True), metrics, sim_summary

# ═══════════════════════════════════════════════════════════════════
# GRID SEARCH OPTIMIZER ENGINE
# ═══════════════════════════════════════════════════════════════════

def optimizer_metrics(out, params):
    """Full metric suite for one quarter's back-test result (`out` from compute_back_test)."""
    actual = out["Actual_Sales"].to_numpy(dtype=float)
    goal   = out["Simulated_Goal"].to_numpy(dtype=float)
    grow   = out["Simulated_Growth"].to_numpy(dtype=float)
    err     = goal - actual
    abs_err = np.abs(err)
    tot_actual = actual.sum()
    n = len(actual)
    with np.errstate(divide="ignore", invalid="ignore"):
        ape = np.where(actual > 0, abs_err / actual, np.nan)
    mean_actual = actual.mean() if n else np.nan
    wmape  = abs_err.sum() / tot_actual if tot_actual else np.nan
    mae    = abs_err.mean() if n else np.nan
    rmse   = float(np.sqrt((err ** 2).mean())) if n else np.nan
    mape   = float(np.nanmean(ape)) if n else np.nan
    medape = float(np.nanmedian(ape)) if n else np.nan
    within = float(np.nanmean((ape <= 0.10).astype(float))) if n else np.nan
    outside = 1 - within if within == within else np.nan
    if n > 1 and goal.std() > 0 and actual.std() > 0:
        corr = float(np.corrcoef(goal, actual)[0, 1])
    else:
        corr = np.nan
    bias = (goal.sum() - actual.sum()) / tot_actual if tot_actual else np.nan
    cap, floor = params["growth_cap"], params["growth_floor"]
    hit = float(np.mean((grow >= cap - 1e-6) | (grow <= floor + 1e-6))) if n else np.nan
    nrmse = rmse / mean_actual if mean_actual else np.nan
    # Composite (lower = better) — governance weights:
    composite = (0.50 * wmape + 0.20 * medape + 0.15 * nrmse +
                 0.10 * outside + 0.05 * hit)
    return {"wmape": wmape, "mae": mae, "rmse": rmse, "mape": mape, "median_ape": medape,
            "pct_within_tol": within, "correlation": corr, "bias": bias,
            "cap_floor_hit_rate": hit, "nrmse": nrmse, "composite": composite}


def evaluate_params_multi_quarter(df, quarter_pairs, params):
    """Back test each (test_q, prior_q) pair; aggregate into a multi-quarter stability score."""
    per_q = []
    for test_q, prior_q in quarter_pairs:
        out, _m, _s = compute_back_test(df, test_q, prior_q, params)
        if out.empty:
            continue
        met = optimizer_metrics(out, params)
        met["test_quarter"] = str(test_q)
        per_q.append(met)
    if not per_q:
        return None
    comp  = np.array([p["composite"] for p in per_q], dtype=float)
    wmape = np.array([p["wmape"]     for p in per_q], dtype=float)
    # Stability penalises combos that win one quarter but swing across others:
    stability = float(np.nanmean(comp) + 0.25 * np.nanstd(wmape))
    return {"avg_composite":    float(np.nanmean(comp)),
            "std_wmape":        float(np.nanstd(wmape)),
            "stability_score":  stability,
            "avg_wmape":        float(np.nanmean(wmape)),
            "avg_mape":         float(np.nanmean([p["mape"]            for p in per_q])),
            "avg_median_ape":   float(np.nanmean([p["median_ape"]     for p in per_q])),
            "avg_rmse":         float(np.nanmean([p["rmse"]           for p in per_q])),
            "avg_within_tol":   float(np.nanmean([p["pct_within_tol"] for p in per_q])),
            "avg_bias":         float(np.nanmean([p["bias"]           for p in per_q])),
            "avg_cap_floor_hit":float(np.nanmean([p["cap_floor_hit_rate"] for p in per_q])),
            "n_quarters":       len(per_q)}


def _frange(lo, hi, step):
    """Inclusive float range, drift-safe."""
    if step <= 0:
        return [round(lo, 6)]
    n = int(round((hi - lo) / step)) + 1
    return [round(lo + i * step, 6) for i in range(max(n, 1))]


def build_quarter_pairs(quarters, max_quarters=4):
    """Consecutive (test, prior) pairs, most recent first, capped to max_quarters."""
    pairs = [(quarters[i], quarters[i - 1]) for i in range(len(quarters) - 1, 0, -1)]
    return pairs[:max_quarters]


def grid_search(df, quarter_pairs, base_params, ranges, progress_cb=None, refine=True):
    """
    ranges: dict of lists for mr_weight(%), vol_balance(%), growth_floor(%), growth_cap(%).
    Coarse pass over `ranges`, then optional fine ±1% pass around the best.
    Returns (results_df sorted by stability_score asc, best_params dict).
    """
    def combos(rng):
        out = []
        for w in rng["mr_weight"]:
            for v in rng["vol_balance"]:
                for f in rng["growth_floor"]:
                    for c in rng["growth_cap"]:
                        if f > c:                       # floor can't exceed cap
                            continue
                        out.append((w, v, f, c))
        return out

    def run(combo_list):
        rows, total = [], len(combo_list)
        for i, (w, v, f, c) in enumerate(combo_list):
            p = dict(base_params)
            p.update({"mr_weight": w / 100.0, "p_weight": 1 - w / 100.0,
                      "vol_balance": v / 100.0,
                      "growth_floor": f / 100.0, "growth_cap": c / 100.0})
            agg = evaluate_params_multi_quarter(df, quarter_pairs, p)
            if agg is not None:
                rows.append({"mr_weight_%": w, "vol_balance_%": v,
                             "growth_floor_%": f, "growth_cap_%": c, **agg})
            if progress_cb:
                progress_cb(i + 1, total)
        return rows

    coarse = run(combos(ranges))
    if not coarse:
        return pd.DataFrame(), None
    df_res = pd.DataFrame(coarse).sort_values("stability_score").reset_index(drop=True)
    best = df_res.iloc[0]

    if refine:
        def around(center, step, lo, hi):
            vals = sorted({round(center + d, 6) for d in (-step, 0, step)})
            return [x for x in vals if lo <= x <= hi]
        fine_rng = {
            "mr_weight":    around(best["mr_weight_%"],    1, 0, 100),
            "vol_balance":  around(best["vol_balance_%"],  1, 0, 20),
            "growth_floor": around(best["growth_floor_%"], 1, -50, 50),
            "growth_cap":   around(best["growth_cap_%"],   1, 0, 100),
        }
        fine = run(combos(fine_rng))
        if fine:
            df_res = (pd.concat([df_res, pd.DataFrame(fine)], ignore_index=True)
                      .drop_duplicates(subset=["mr_weight_%", "vol_balance_%",
                                               "growth_floor_%", "growth_cap_%"])
                      .sort_values("stability_score").reset_index(drop=True))
            best = df_res.iloc[0]

    best_params = dict(base_params)
    best_params.update({"mr_weight": best["mr_weight_%"] / 100.0,
                        "p_weight": 1 - best["mr_weight_%"] / 100.0,
                        "vol_balance": best["vol_balance_%"] / 100.0,
                        "growth_floor": best["growth_floor_%"] / 100.0,
                        "growth_cap": best["growth_cap_%"] / 100.0})
    return df_res, best_params

# ═══════════════════════════════════════════════════════════════════
# METRIC CARD PRESETS
# ═══════════════════════════════════════════════════════════════════
if is_dark:
    LIGHT_CARDS = dict(background_color="#1e293b", border_left_color="#2563eb",
                       border_color="#334155", box_shadow=True)
    DARK_CARDS  = dict(background_color="#0f172a", border_left_color="#2563eb",
                       border_color="#1e293b", box_shadow=True)
else:
    LIGHT_CARDS = dict(background_color="#f0f7ff", border_left_color="#2563eb",
                       border_color="#dbeafe", box_shadow=True)
    DARK_CARDS  = dict(background_color="#eff6ff", border_left_color="#2563eb",
                       border_color="#bfdbfe", box_shadow=True)

# ═══════════════════════════════════════════════════════════════════
# TAB 0 — INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════════
if active == 0:
    st.markdown("""
    <div class="instructions-hero">
        <h1>📖 Welcome to the Goal Setting Tool</h1>
        <p>A platform for setting national sales goals and allocating them across
           territories using fair-share, equal, or weighted models. Read the steps
           below, then click <strong>Start</strong> to begin.</p>
    </div>
    """, unsafe_allow_html=True)

    steps = [
        'Upload the input data in the correct format <em>(template provided on the "Input & Validation" tab).</em>',
        'Verify the data quality summary - fix any missing values, duplicates, or outliers flagged by the tool before proceeding.',
        'Set the date range for the period you want to analyse, then review the territory performance summary.',
        'On the <strong>National Goal Setting</strong> tab, enter the total national sales goal for the upcoming period.',
        'Choose an allocation model - Fair Share, Equal, or Weighted - to distribute the national goal across territories.',
        'Use <strong>Back Testing</strong> to validate your goal model against prior periods before finalising.',
        'View and download the final territory goals from the <strong>Final Allocation</strong> tab as an Excel file.',
        'If your dataset is very large (millions of rows), make sure your system has enough memory before processing.',
    ]
    steps_html = '<div class="steps-list">'
    for i, step in enumerate(steps, 1):
        steps_html += (
            f'<div class="step-item">'
            f'<div class="step-num">{i}</div>'
            f'<div class="step-text">{step}</div>'
            f'</div>'
        )
    steps_html += '</div>'
    st.markdown(steps_html, unsafe_allow_html=True)

    _, btn_col, _ = st.columns([1, 1, 1])
    with btn_col:
        if st.button("Start", type="primary", use_container_width=True, key="start_btn"):
            go_to_tab(1)


# ═══════════════════════════════════════════════════════════════════
# TAB 1 — INPUT & VALIDATION
# ═══════════════════════════════════════════════════════════════════
elif active == 1:
    st.subheader("Upload Data")
    st.caption("Required columns: Week · Territory ID · Territory Name · Product · Sales · Goals")

    file = st.file_uploader("Upload CSV or Excel file", type=["csv", "xlsx"])

    if not file:
        if st.session_state.data is None:
            st.info("📁 Please upload a CSV or Excel (.xlsx) file to get started.")
            st.stop()
    else:
        if st.session_state.data is None:
            raw = preprocess(load_file(file))
            missing_cols = validate_schema(raw)
            if missing_cols:
                st.error(f"❌ Missing columns: **{', '.join(missing_cols)}**")
                st.stop()
            if raw["Week"].isnull().sum() > 0:
                st.error("❌ Could not parse **Week** column — check date format.")
                st.stop()
            st.session_state.raw_data = raw.copy()
            st.session_state.data     = raw.copy()

    data = st.session_state.data
    if data is None:
        st.stop()

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Rows",  f"{len(data):,}")
    c2.metric("Territories", data["Territory ID"].nunique())
    c3.metric("Weeks",       data["Week"].nunique())
    style_metric_cards(**LIGHT_CARDS)

    st.divider()

    # ── DATE RANGE FILTER ────────────────────────────────────────
    st.subheader("Date Range Filter")
    min_d, max_d = data["Week"].min().date(), data["Week"].max().date()
    dc1, dc2 = st.columns(2)
    start = dc1.date_input("From", min_d, min_value=min_d, max_value=max_d, key="d_start")
    end   = dc2.date_input("To",   max_d, min_value=min_d, max_value=max_d, key="d_end")

    if start > end:
        st.error("Start date must be before end date.")
        st.stop()

    date_filtered = apply_date_filter(data, start, end)

    st.divider()

    # ── PRODUCT SELECTION ────────────────────────────────────────
    st.subheader("Product Selection")
    st.caption("Pick which product(s) to analyse. Everything below — data quality, summary, "
               "and downloads — respects this selection and the date range above.")

    all_products = sorted(data["Product"].dropna().astype(str).unique().tolist())

    with st.expander("🧪 Select Product(s)", expanded=True):
        select_all = st.checkbox("✅ All Products (Select All)", value=True,
                                 key="prod_select_all")

        chosen = []
        n_cols = 3
        cols = st.columns(n_cols)
        for i, prod in enumerate(all_products):
            with cols[i % n_cols]:
                checked = st.checkbox(prod, value=select_all,
                                      key=f"prod_cb_{i}", disabled=select_all)
                if checked:
                    chosen.append(prod)

        if select_all:
            selected_products = all_products
        else:
            selected_products = chosen if chosen else all_products
            if not chosen:
                st.warning("No product ticked — showing **all products** until you pick at least one.")

    st.session_state.selected_products = selected_products

    # Apply BOTH date + product filters → working scope for the rest of the page
    filtered = apply_product_filter(date_filtered, selected_products)
    st.session_state.filtered_data = filtered.copy()

    prod_label = "All products" if select_all else ", ".join(selected_products)
    st.success(
        f"✅ **{len(filtered):,} rows** in scope — {prod_label} · "
        f"{start.strftime('%d %b %Y')} → {end.strftime('%d %b %Y')}"
    )

    # ── PRODUCT STATISTICS ───────────────────────────────────────
    st.subheader("Product Statistics")
    st.caption("Per-product and combined — within the selected date range.")

    def _prod_stats_block(d):
        if d.empty:
            st.info("No rows for this selection.")
            return
        ts, tg = d["Sales"].sum(), d["Goals"].sum()
        att = (ts / tg * 100) if tg > 0 else 0
        a1, a2, a3 = st.columns(3)
        a1.metric("Average Sales", fmt(d["Sales"].mean()))
        a2.metric("Maximum Sales", fmt(d["Sales"].max()))
        a3.metric("Minimum Sales", fmt(d["Sales"].min()))
        b1, b2, b3 = st.columns(3)
        b1.metric("Total Sales", fmt(ts))
        b2.metric("Total Goals", fmt(tg))
        b3.metric("Attainment",  f"{att:.1f}%")
        style_metric_cards(**DARK_CARDS)

    if len(selected_products) > 1:
        stat_scope = st.radio("Statistics for",
                              ["Combined"] + selected_products,
                              horizontal=True, key="prod_stat_scope")
        if stat_scope == "Combined":
            _prod_stats_block(filtered)
        else:
            _prod_stats_block(filtered[filtered["Product"] == stat_scope])
    else:
        _prod_stats_block(filtered)

    # ── DATA QUALITY ─────────────────────────────────────────────
    st.subheader("Data Quality")
    st.markdown(
        "Below are three checks for data integrity — **missing values**, **duplicate rows**, "
        "and **statistical outliers** — all computed on the selected product(s) and date range. "
        "The fix is only applied when you click **Apply Fix**."
    )
    work = filtered.copy()

    null_df  = work[work.isnull().any(axis=1)]
    null_cnt = len(null_df)
    dup_df   = work[work.duplicated()]
    dup_cnt  = len(dup_df)

    # Product-wise outlier z-scores (each row vs its OWN product's mean/std)
    z_scores = product_wise_zscores(work)
    out_cnt  = int((z_scores > 3).sum()) if not z_scores.empty else 0

    # ── Size-aware data quality score ─────────────────────────────
    total_rows = max(len(work), 1)
    null_rows  = int(work.isnull().any(axis=1).sum())
    null_pct   = null_rows / total_rows
    dup_pct    = dup_cnt   / total_rows
    out_pct    = out_cnt   / total_rows

    score = 100 - (null_pct * 40 + dup_pct * 30 + out_pct * 30)
    score = int(round(max(0, min(100, score))))

    qc1, qc2, qc3, qc4 = st.columns(4)
    qc1.metric("Missing Values", null_cnt,
               delta="Clean" if null_cnt == 0 else "Found",
               delta_color="normal" if null_cnt == 0 else "inverse")
    qc2.metric("Duplicate Rows", dup_cnt,
               delta="Clean" if dup_cnt == 0 else "Found",
               delta_color="normal" if dup_cnt == 0 else "inverse")
    qc3.metric("Outliers (±3σ)", out_cnt,
               delta="Clean" if out_cnt == 0 else "Review",
               delta_color="normal" if out_cnt == 0 else "off")
    qc4.metric("Quality Score",  f"{score}/100")
    style_metric_cards(**LIGHT_CARDS)
    st.progress(score / 100)

    qc_axis_color  = "#f1f5f9" if is_dark else "#0f172a"
    qc_grid_color  = "#1f2937" if is_dark else "#e2e8f0"
    qc_tooltip_bg  = "#1e293b" if is_dark else "#ffffff"
    qc_tooltip_fg  = "#f1f5f9" if is_dark else "#0f172a"
    qc_tooltip_brd = "#334155" if is_dark else "#cbd5e1"

    # ════════════════════════════════════════════════════════════
    # MISSING VALUES (product-wise fill + manual edit + preview)
    # ════════════════════════════════════════════════════════════
    if null_cnt > 0:
        with st.expander(f"🔧 Fix {null_cnt} missing value row(s)", expanded=True):
            st.markdown(
                f"**{null_cnt}** row(s) have at least one missing value. **Mean / Median / Mode "
                "are calculated per product** — each product's gaps are filled using only that "
                "product's own values."
            )

            st.markdown("**Affected rows (preview):**")
            st.dataframe(null_df.head(20), use_container_width=True, hide_index=True)

            null_action_help = {
                "Do Nothing":                "Keep the rows exactly as they are.",
                "Fill with 0":               "Replace every missing value with 0.",
                "Fill with Mean":            "Fill missing Sales/Goals with that product's average.",
                "Fill with Median":          "Fill missing Sales/Goals with that product's median.",
                "Fill with Mode":            "Fill missing Sales/Goals with that product's most frequent value.",
                "Fill using Previous Value": "Carry the previous (earlier-week) value forward, per product.",
                "Fill using Next Value":     "Carry the next (later-week) value backward, per product.",
                "Remove Rows":               "Drop any row (in scope) that has a missing value.",
            }
            action = st.radio("Action", list(null_action_help.keys()),
                              key="null_act", horizontal=True)
            st.caption(null_action_help[action])

            # ── Manual editing when fewer than 10 missing rows ──
            manual_edit = False
            edited_rows = None
            if null_cnt < 10:
                manual_edit = st.checkbox(
                    "✏️ Edit these values manually instead",
                    key="null_manual",
                    help="Because there are fewer than 10 missing rows, you can type the "
                         "correct Sales / Goals values directly.")
                if manual_edit:
                    st.caption("Edit the Sales / Goals cells below, then click **Apply Fix**.")
                    edited_rows = st.data_editor(
                        null_df.copy(),
                        use_container_width=True, hide_index=False,
                        key="null_editor",
                        disabled=["Week", "Territory ID", "Territory Name", "Product"],
                    )

            if st.button("Apply Fix", key="null_apply", type="primary"):
                full = st.session_state.data.copy()

                if manual_edit and edited_rows is not None:
                    full.loc[edited_rows.index, ["Sales", "Goals"]] = \
                        edited_rows[["Sales", "Goals"]]
                    msg = "✅ Missing values updated manually."
                elif action == "Remove Rows":
                    drop_idx = null_df.index  # only in-scope null rows
                    full = full.drop(index=drop_idx)
                    msg = f"✅ {len(drop_idx):,} row(s) with missing values removed."
                elif action == "Do Nothing":
                    msg = "ℹ️ No changes applied (you chose *Do Nothing*)."
                else:
                    full = fill_missing_product_wise(full, selected_products, action)
                    msg = f"✅ Missing values handled — **{action}** (per product)."

                st.session_state.null_fixed_idx = null_df.index.tolist()
                st.session_state.data           = full
                st.session_state.action_type    = "missing"
                st.session_state.action_msg     = msg
                st.rerun()

    elif st.session_state.action_type == "missing":
        st.success(st.session_state.action_msg or "✅ Missing values resolved.")
        fixed_idx = st.session_state.get("null_fixed_idx", [])
        with st.expander("🔍 Preview — previously-missing rows after fix", expanded=True):
            avail = [i for i in fixed_idx if i in st.session_state.data.index]
            if avail:
                st.caption("These rows had missing values — now showing their updated values.")
                st.dataframe(st.session_state.data.loc[avail].head(20),
                             use_container_width=True)
            else:
                st.caption("Those rows were removed from the dataset.")
        st.session_state.action_type = None
    else:
        st.success("✅ No missing values found.")

    # ════════════════════════════════════════════════════════════
    # DUPLICATE ROWS
    # ════════════════════════════════════════════════════════════
    if dup_cnt > 0:
        with st.expander(f"🔧 Fix {dup_cnt} duplicate row(s)", expanded=True):
            st.markdown(
                f"**{dup_cnt}** row(s) appear more than once. Duplicate rows inflate totals — "
                "removing them is almost always correct."
            )
            st.markdown("**Duplicate rows (preview):**")
            st.dataframe(dup_df.head(20), use_container_width=True, hide_index=True)
            st.caption("Only second-and-later occurrences are shown. The first occurrence is kept.")

            if st.button("Apply Fix — Remove Duplicates", key="dup_apply", type="primary"):
                full    = st.session_state.data.drop_duplicates()
                removed = len(st.session_state.data) - len(full)
                st.session_state.data        = full
                st.session_state.action_type = "duplicate"
                st.session_state.action_msg  = f"✅ {removed:,} duplicate row(s) removed successfully."
                st.rerun()

    elif st.session_state.action_type == "duplicate":
        st.success(st.session_state.action_msg or "✅ Duplicates removed successfully.")
        with st.expander("✅ View updated data preview (after fix)", expanded=False):
            updated = apply_product_filter(
                apply_date_filter(st.session_state.data, start, end), selected_products)
            st.caption(f"Showing first 20 rows of the **updated** dataset "
                       f"({len(updated):,} rows in scope).")
            st.dataframe(updated.head(20), use_container_width=True, hide_index=True)
        st.session_state.action_type = None
    else:
        st.success("✅ No duplicate rows found.")

    # ════════════════════════════════════════════════════════════
    # OUTLIERS (product-wise + scope radio)
    # ════════════════════════════════════════════════════════════
    if out_cnt > 0:
        with st.expander(f"🔧 Review {out_cnt} outlier row(s)", expanded=True):
            st.markdown(
                f"**{out_cnt}** sales values are statistical outliers — more than **3 standard "
                "deviations** from their **product's** mean. Outliers can be legitimate "
                "(a big order, a launch month) or data-entry errors."
            )

            # Radio only when more than one product is selected
            if len(selected_products) > 1:
                out_scope = st.radio("Show outliers for",
                                     ["All"] + selected_products,
                                     horizontal=True, key="out_scope")
            else:
                out_scope = "All"

            if out_scope == "All":
                scope_df = work.copy()
            else:
                scope_df = work[work["Product"] == out_scope].copy()

            scope_z   = product_wise_zscores(scope_df)
            scope_out = scope_df[scope_z > 3].copy()
            scope_out_cnt = len(scope_out)

            sales_mean = scope_df["Sales"].mean()
            sales_std  = scope_df["Sales"].std()
            upper_cap  = sales_mean + 3 * sales_std
            lower_cap  = sales_mean - 3 * sales_std

            os1, os2, os3, os4 = st.columns(4)
            os1.metric("Mean Sales",        fmt(sales_mean))
            os2.metric("Std Dev",           fmt(sales_std))
            os3.metric("Upper Bound (+3σ)", fmt(upper_cap))
            os4.metric("Outlier Rows",      f"{scope_out_cnt:,}")

            single_scope = (out_scope != "All") or (len(selected_products) == 1)

            st.markdown("**Outlier distribution chart**")
            st.caption(
                "Each dot is a row. Red diamonds are flagged outliers (beyond ±3σ within their "
                "product)." + (" The dashed line is the +3σ bound." if single_scope else
                "")
            )

            chart_df = scope_df.reset_index(drop=True).copy()
            chart_df["row_idx"]    = chart_df.index
            chart_df["is_outlier"] = (scope_z.reset_index(drop=True) > 3)
            chart_df["label"]      = (chart_df["Territory Name"].astype(str) + " · " +
                                       chart_df["Product"].astype(str) + " · " +
                                       chart_df["Week"].dt.strftime("%Y-%m-%d"))

            normal_pts  = chart_df[~chart_df["is_outlier"]]
            outlier_pts = chart_df[chart_df["is_outlier"]]

            fig_out = go.Figure()
            fig_out.add_trace(go.Scatter(
                x=normal_pts["row_idx"], y=normal_pts["Sales"],
                mode="markers", name="Normal",
                marker=dict(size=5, color="#2563eb", opacity=0.55, line=dict(width=0)),
                customdata=normal_pts["label"],
                hovertemplate="<b>%{customdata}</b><br>Sales: %{y:,.0f}<extra></extra>",
            ))
            fig_out.add_trace(go.Scatter(
                x=outlier_pts["row_idx"], y=outlier_pts["Sales"],
                mode="markers", name="Outlier (>3σ)",
                marker=dict(size=11, color="#ef4444", opacity=0.9,
                            line=dict(width=1.5, color="white" if is_dark else "#7f1d1d"),
                            symbol="diamond"),
                customdata=outlier_pts["label"],
                hovertemplate="<b>%{customdata}</b><br>Sales: %{y:,.0f} ⚠️<extra></extra>",
            ))
            if single_scope:
                fig_out.add_hline(y=upper_cap, line_dash="dash", line_color="#f97316",
                                  line_width=1.5,
                                  annotation_text=f"+3σ ({upper_cap:,.0f})",
                                  annotation_position="top right",
                                  annotation_font_color=qc_axis_color)
                fig_out.add_hline(y=sales_mean, line_dash="dot", line_color="#64748b",
                                  line_width=1,
                                  annotation_text=f"Mean ({sales_mean:,.0f})",
                                  annotation_position="top left",
                                  annotation_font_color=qc_axis_color)
                if lower_cap > 0:
                    fig_out.add_hline(y=lower_cap, line_dash="dash", line_color="#f97316",
                                      line_width=1.5,
                                      annotation_text=f"−3σ ({lower_cap:,.0f})",
                                      annotation_position="bottom right",
                                      annotation_font_color=qc_axis_color)
            fig_out.update_layout(
                template="plotly_dark" if is_dark else "plotly_white",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                height=380, margin=dict(l=70, r=20, t=30, b=70),
                legend=dict(orientation="h", y=1.08, x=1, xanchor="right",
                            font=dict(color=qc_axis_color)),
                xaxis=dict(
                    title=dict(text="Row Index",
                               font=dict(color=qc_axis_color, size=14, family="Calibri"),
                               standoff=12),
                    color=qc_axis_color, tickfont=dict(color=qc_axis_color, size=11),
                    gridcolor=qc_grid_color, linecolor=qc_grid_color),
                yaxis=dict(
                    title=dict(text="Sales ($)",
                               font=dict(color=qc_axis_color, size=14, family="Calibri"),
                               standoff=12),
                    tickformat=",.0f", color=qc_axis_color,
                    tickfont=dict(color=qc_axis_color, size=11),
                    gridcolor=qc_grid_color, linecolor=qc_grid_color),
                font=dict(color=qc_axis_color),
                hoverlabel=dict(bgcolor=qc_tooltip_bg, bordercolor=qc_tooltip_brd,
                                font=dict(color=qc_tooltip_fg, size=12)),
            )
            st.plotly_chart(fig_out, use_container_width=True)

            st.markdown("**Outlier rows (preview):**")
            out_rows = scope_out.copy()
            out_rows["Z-Score"] = scope_z[scope_z > 3].round(2).values
            st.dataframe(out_rows.head(20), use_container_width=True, hide_index=True)

            oa_help = {
                "Keep as-is":                     "Leave outlier values untouched.",
                "Cap at 3 Std Dev (per product)": "Cap each selected product's Sales at its own +3σ.",
            }
            oa = st.radio("Action", list(oa_help.keys()), key="out_act", horizontal=True)
            st.caption(oa_help[oa])

            if st.button("Apply Fix", key="out_apply", type="primary"):
                full = st.session_state.data.copy()
                if oa.startswith("Cap"):
                    full, n_capped = cap_outliers_product_wise(full, selected_products)
                    st.session_state.action_msg = (
                        f"✅ Outliers capped per product at +3σ. {n_capped:,} value(s) adjusted.")
                else:
                    st.session_state.action_msg = "ℹ️ Outliers kept as-is (no changes applied)."
                st.session_state.data        = full
                st.session_state.action_type = "outlier"
                st.rerun()

    elif st.session_state.action_type == "outlier":
        if "capped" in (st.session_state.action_msg or "").lower():
            st.success(st.session_state.action_msg)
        else:
            st.info(st.session_state.action_msg or "Outliers handled.")
        with st.expander("✅ View updated data preview (after fix)", expanded=False):
            updated = apply_product_filter(
                apply_date_filter(st.session_state.data, start, end), selected_products)
            st.caption(f"Showing first 20 rows of the **updated** dataset "
                       f"({len(updated):,} rows in scope). Sales reflects any cap applied.")
            st.dataframe(updated.head(20), use_container_width=True, hide_index=True)
        st.session_state.action_type = None
    else:
        st.success("✅ No outliers detected.")

    st.divider()

    # ── SUMMARY (product radio scope) ────────────────────────────
    st.subheader("Summary")
    if len(selected_products) > 1:
        summary_scope = st.radio("View summary for",
                                 ["All"] + selected_products,
                                 horizontal=True, key="summary_scope")
    else:
        summary_scope = "All"

    # ALL selected products + date filter — used for the DOWNLOAD (ignores the radio)
    clean_scope = apply_product_filter(
        apply_date_filter(st.session_state.data, start, end), selected_products)

    # Summary VIEW scope — single product if the radio picked one
    if summary_scope == "All":
        plot_df = clean_scope.copy()
    else:
        plot_df = clean_scope[clean_scope["Product"] == summary_scope].copy()

    if plot_df.empty:
        st.info("No rows for this product selection.")
    else:
        total_sales = plot_df["Sales"].sum()
        total_goals = plot_df["Goals"].sum()
        attainment  = (total_sales / total_goals * 100) if total_goals > 0 else 0

        # Quarter-over-Quarter growth (last quarter vs the quarter before it)
        q_sales = (plot_df.assign(_Q=plot_df["Week"].dt.to_period("Q"))
                          .groupby("_Q")["Sales"].sum().sort_index())
        if len(q_sales) >= 2 and q_sales.iloc[-2] != 0:
            qoq = (q_sales.iloc[-1] - q_sales.iloc[-2]) / q_sales.iloc[-2] * 100
        else:
            qoq = 0

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Sales",  fmt(total_sales))
        k2.metric("Total Goals",  fmt(total_goals))
        k3.metric("Attainment",   f"{attainment:.1f}%")
        k4.metric("QoQ Growth",   f"{qoq:+.1f}%")
        style_metric_cards(**DARK_CARDS)

        st.divider()

        st.subheader("Sales vs Goals Trend")
        level = st.selectbox("Aggregation", ["Monthly", "Weekly", "Quarterly"], key="trend_lvl")
        st.plotly_chart(plot_trend(plot_df, level), use_container_width=True)

        st.divider()

        st.subheader("Territory Performance")
        terr = (plot_df.groupby(["Territory ID", "Territory Name"])[["Sales", "Goals"]]
                .sum().reset_index().sort_values("Sales", ascending=False))
        terr["Attainment %"] = (terr["Sales"] / terr["Goals"] * 100).round(1)

        tc1, tc2 = st.columns(2)
        with tc1:
            st.markdown("**🥇 Top 5 Territories**")
            st.dataframe(
                terr.head(5)[["Territory Name", "Sales", "Attainment %"]]
                .style.format({"Sales": "{:,.0f}", "Attainment %": "{:.1f}%"}),
                use_container_width=True, hide_index=True)
        with tc2:
            st.markdown("**⚠️ Bottom 5 Territories**")
            st.dataframe(
                terr.tail(5)[["Territory Name", "Sales", "Attainment %"]]
                .style.format({"Sales": "{:,.0f}", "Attainment %": "{:.1f}%"}),
                use_container_width=True, hide_index=True)

        with st.expander("📄 Raw Data Preview (first 100 rows)"):
            st.dataframe(plot_df.head(100), use_container_width=True, hide_index=True)

    st.divider()

    # ── DOWNLOAD + NEXT ──────────────────────────────────────────
    dl_col, nxt_col = st.columns(2)
    with dl_col:
        st.download_button(
            "⬇️ Download Cleaned Data",
            data=to_excel(clean_scope, "Clean_Data",
                          currency_cols=["Sales", "Goals"],
                          title="Cleaned Sales Data"),
            file_name="clean_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with nxt_col:
        if st.button("Next: National Goal Setting →", use_container_width=True, type="primary"):
            go_to_tab(2)


# ═══════════════════════════════════════════════════════════════════
# TAB 2 — NATIONAL GOAL SETTING
# ═══════════════════════════════════════════════════════════════════
elif active == 2:

    if st.session_state.data is None:
        st.warning("⚠️ Please upload and validate your data first.")
        if st.button("← Back to Input & Validation"):
            go_to_tab(1)
        st.stop()

    full_df  = st.session_state.data.copy()
    products = st.session_state.selected_products or sorted(
        full_df["Product"].dropna().astype(str).unique().tolist())

    def _reset_product_state():
        for k in ("nation_goal", "weighted_result_df", "chosen_params",
                  "bt_result_df", "bt_metrics", "opt_results_df", "opt_best_params"):
            st.session_state[k] = None
        st.session_state.nation_goal_submitted = False
        st.session_state.ngs_model_tab = None
        st.session_state.bt_intro_seen = False
        st.session_state.bt_mode = None
        st.session_state.chosen_params_src = None

    # ── Product picker (only when more than one product was selected) ──
    if len(products) > 1 and st.session_state.gs_product is None:
        st.subheader("Choose Product for National Goal Setting")
        st.caption("Goal setting runs one product at a time. Pick a product to begin.")
        done = st.session_state.gs_completed or []
        if done:
            st.success("✅ Completed so far: " + ", ".join(done))
        remaining = [p for p in products if p not in done] or products
        pick = st.radio("Product", remaining, key="gs_pick")
        if st.button("Submit", type="primary"):
            st.session_state.gs_product = pick
            _reset_product_state()
            st.rerun()
        st.stop()

    # Single product → auto-select, skip the picker
    if st.session_state.gs_product is None:
        st.session_state.gs_product = products[0]

    gs_product = st.session_state.gs_product
    df     = full_df[full_df["Product"] == gs_product].copy()
    n_terr = df["Territory ID"].nunique()

    st.caption(f"🧪 Goal setting for product: **{gs_product}**")
    
    if not st.session_state.nation_goal_submitted:
        st.subheader("Enter National Goal")
        st.info(
            "💡 Enter the total national sales goal for the goal period. "
            "After submission, choose from three allocation models to distribute across territories."
        )

        ng_col, _ = st.columns([2, 2])
        with ng_col:
            ng_input = st.number_input(
                "National Goal (total across all territories)",
                min_value=0.0,
                value=float(st.session_state.nation_goal) if st.session_state.nation_goal else None,
                step=1000.0,
                format="%.0f",
                placeholder="Enter national goal (e.g. 5000000)",
            )
            # Treat empty input as 0 for downstream logic
            if ng_input is None:
                ng_input = 0.0

        if ng_input > 0:
            df_t = df.copy()
            df_t["Month"] = df_t["Week"].dt.to_period("M")
            mr3_m          = sorted(df_t["Month"].unique())[-3:]
            total_mr3_prev = df_t[df_t["Month"].isin(mr3_m)]["Sales"].sum()
            ig  = (ng_input / total_mr3_prev - 1) if total_mr3_prev > 0 else 0
            ppa = ng_input / n_terr if n_terr > 0 else 0

            st.divider()
            p1, p2, p3 = st.columns(3)
            p1.metric("Nation Goal",       fmt(ng_input))
            p2.metric("Implied Growth",    pct_fmt(ig),
                      delta="vs MR3 baseline",
                      delta_color="normal" if ig >= 0 else "inverse")
            p3.metric("Per-Territory Avg", fmt(ppa),
                      delta=f"across {n_terr} territories", delta_color="off")
            style_metric_cards(**DARK_CARDS)

        st.divider()
        if st.button("Submit & View Allocation Models →", type="primary"):
            if ng_input <= 0:
                st.error("Please enter a value greater than 0.")
            else:
                st.session_state.nation_goal           = ng_input
                st.session_state.nation_goal_submitted = True
                st.rerun()

    else:
        nation_goal = st.session_state.nation_goal

        df_h = df.copy()
        df_h["Month"] = df_h["Week"].dt.to_period("M")
        mr3_h        = sorted(df_h["Month"].unique())[-3:]
        total_mr3_h  = df_h[df_h["Month"].isin(mr3_h)]["Sales"].sum()
        ig_h         = (nation_goal / total_mr3_h - 1) if total_mr3_h > 0 else 0

        st.subheader("National Goal")
        h1, h2, h3 = st.columns(3)
        h1.metric("Nation Goal",   fmt(nation_goal))
        h2.metric("Growth vs MR3", pct_fmt(ig_h))
        h3.metric("Territories",   n_terr)
        style_metric_cards(**DARK_CARDS)

        if st.button("✏️ Edit Goal"):
            st.session_state.nation_goal_submitted = False
            st.rerun()

        st.divider()

        st.subheader("Select Allocation Model")
        model_names = ["Weighted Model", "Fair Share Model", "Equal Allocation Model"]
        mc1, mc2, mc3 = st.columns(3)
        for i, (col, name) in enumerate(zip([mc1, mc2, mc3], model_names)):
            with col:
                if st.button(name, key=f"mt_{i}", use_container_width=True):
                    st.session_state.ngs_model_tab = i
                    st.rerun()

        selected = st.session_state.ngs_model_tab
        if selected is None:
            st.stop()

        st.caption(f"Active: **{model_names[selected]}**")
        st.divider()

        # ── WEIGHTED MODEL ────────────────────────────────────────
        if selected == 0:
            st.subheader("Weighted Model")
            st.markdown(
                "Distributes the National Goal across territories using a **blend of recent and "
                "prior performance**, with safeguards to prevent extreme growth targets. "
                "Each territory's goal is built up in stages so the impact of every parameter "
                "is transparent."
            )

            # ── Parameters ────────────────────────────────────────
            with st.expander("⚙️  Model Parameters", expanded=True):
                st.markdown(
                    "Tune how the model balances recent vs prior sales, smooths out volume "
                    "differences, and limits how aggressive growth targets can be. "
                    "The defaults below are a good starting point for most products."
                )

                # ─ Row 1: Weights & scaling ─
                st.markdown(
                    "**MR Weight** controls how much weight the *Most Recent* period gets in the blended "
                    "goal calculation. **P Weight** is automatically `100% − MR Weight`. "
                    "**Historical Sales Scale** scales the raw historical sales up or down before any "
                    "calculation begins (leave at 100% unless adjusting for known seasonality or data issues)."
                )
                pc1, pc2, pc3 = st.columns(3)
                with pc1:
                    mr_weight_pct = st.number_input(
                        "Baseline Weight (MR) %",
                        min_value=0.0, max_value=100.0, value=70.0, step=0.5,
                        format="%.1f",
                        help="Weight applied to Most Recent period sales. Decimals OK.",
                    )
                    mr_weight = mr_weight_pct / 100.0
                with pc2:
                    p_weight = 1.0 - mr_weight
                    st.metric("P Weight (Prior)", f"{p_weight * 100:.0f}%",
                              delta="auto = 100% − MR Weight", delta_color="off")
                with pc3:
                    hist_scale_pct = st.number_input(
                        "Historical Sales Scale %",
                        min_value=0.0, max_value=200.0, value=100.0, step=0.5,
                        format="%.1f",
                        help="Scaling factor applied to all historical sales.",
                    )
                    hist_scale = hist_scale_pct / 100.0

                # ─ Row 2: Smoothing + cap/floor ─
                st.markdown(
                    "**Volume Balance** is a smoothing factor that gently pulls small territories up and "
                    "large territories down toward the average. **Growth Cap** is the maximum allowed "
                    "growth % any single territory can be asked to deliver. **Growth Floor** is the minimum — "
                    "use it to ensure no territory is given a declining goal."
                )
                pc4, pc5, pc6 = st.columns(3)
                with pc4:
                    # Volume Balance shown in basis points × 10 (0.0%–10.0%, step 0.5%)
                    vol_balance_pct = st.number_input(
                        "Volume Balance %",
                        min_value=0.0, max_value=10.0, value=1.5, step=0.1,
                        format="%.1f",
                        help="Volume balancing moderator. Higher = stronger smoothing.",
                    )
                    vol_balance = vol_balance_pct / 100.0
                with pc5:
                    growth_cap_pct = st.number_input(
                        "Growth Cap (Max) %",
                        min_value=0.0, max_value=100.0, value=6.0, step=0.5,
                        format="%.1f",
                        help="Maximum allowed growth % per territory.",
                    )
                    growth_cap = growth_cap_pct / 100.0
                with pc6:
                    growth_floor_pct = st.number_input(
                        "Growth Floor (Min) %",
                        min_value=-50.0, max_value=50.0, value=0.0, step=0.5,
                        format="%.1f",
                        help="Minimum allowed growth % per territory.",
                    )
                    growth_floor = growth_floor_pct / 100.0

                # ─ Row 3: Flat goal + windows ─
                st.markdown(
                    "**Flat Goal %** is the portion of the National Goal split equally across all "
                    "territories *before* the weighted blend is applied (set to 0% for a fully "
                    "performance-based allocation). **MR Window** is how many recent months count as "
                    "*Most Recent* (default 3). **P Window** is how many months before that count as "
                    "*Prior* (default 3)."
                )
                pc7, pc8, pc9 = st.columns(3)
                with pc7:
                    flat_goal_pct_int = st.number_input(
                        "Flat Goal %",
                        min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                        format="%.1f",
                        help="Equal-share portion of National Goal before weighting.",
                    )
                    flat_goal_pct = flat_goal_pct_int / 100.0
                with pc8:
                    mr3_window = st.number_input(
                        "MR Window (months)", min_value=1, max_value=12,
                        value=3, step=1,
                        help="Number of most-recent months for the MR baseline.",
                    )
                with pc9:
                    p3_window = st.number_input(
                        "P Window (months)", min_value=1, max_value=12,
                        value=3, step=1,
                        help="Number of prior months for the P baseline.",
                    )

            params = {
                "hist_scale":    hist_scale,
                "flat_goal_pct": flat_goal_pct,
                "mr_weight":     mr_weight,
                "p_weight":      p_weight,
                "vol_balance":   vol_balance,
                "growth_cap":    growth_cap,
                "growth_floor":  growth_floor,
                "mr3_window":    int(mr3_window),
                "p3_window":     int(p3_window),
            }

            result, summary = compute_weighted_model(df, nation_goal, params)

            # Persist results so the Final Allocation tab can show them
            st.session_state.weighted_result_df    = result
            st.session_state.weighted_summary      = summary
            st.session_state.weighted_params       = params
            st.session_state.selected_model_label  = "Weighted Model"

            st.divider()

            # ── Reconciliation banner ─────────────────────────────
            diff = summary["final_minus_nation"]
            if summary.get("fallback_triggered"):
                eff_cap   = summary["effective_cap"]   * 100
                eff_floor = summary["effective_floor"] * 100
                user_cap  = summary["user_cap"]        * 100
                user_floor= summary["user_floor"]      * 100
                st.warning(
                    f"⚠️ **Cap/floor limits had to be relaxed to fully allocate the National Goal.**  \n"
                    f"Your settings: cap **{user_cap:.1f}%**, floor **{user_floor:.1f}%**. "
                    f"Effective range after pro-rata fallback: **{eff_floor:.1f}% to {eff_cap:.1f}%**.  \n"
                    f"This happens when the National Goal demands more (or less) growth than your "
                    f"cap/floor allows for *every* territory. The leftover was spread proportionally "
                    f"across all territories so allocation matches the National Goal exactly. "
                    f"Consider widening the cap/floor or lowering the National Goal to avoid this."
                )
            elif abs(diff) < 1:
                st.success(f"✅ Final allocated total matches the National Goal "
                           f"(difference = {diff:+,.2f}).")
            else:
                st.warning(
                    f"⚠️ Final allocated total differs from the National Goal by **{diff:+,.2f}**. "
                    "Check parameter settings if this is unexpected."
                )

            # ── Headline KPIs ─────────────────────────────────────
            st.subheader("Summary")
            st.markdown(
                "High-level totals showing the National Goal alongside the **Total Allocated** "
                "(sum of all territory goals). The **MR Sales** column is total sales from the "
                "most recent period (3 months by default), and **P Sales** is total sales from "
                "the prior period (also 3 months by default). Both serve as the baseline against "
                "which growth is measured."
            )
            h1, h2, h3, h4 = st.columns(4)
            h1.metric("National Goal",       fmt(nation_goal))
            h2.metric("Total Allocated",     fmt(summary["final_total_goal"]))
            h3.metric(f"MR{params['mr3_window']} Sales", fmt(summary["total_mr3"]))
            h4.metric(f"P{params['p3_window']} Sales",   fmt(summary["total_p3"]))
            style_metric_cards(**DARK_CARDS)

            # ── Cap / Floor Adjustment Summary ────────────────────
            st.subheader("Adjustment & Redistribution")
            st.markdown(
                "When some territories hit the Growth Cap or Growth Floor, their excess (or shortfall) "
                "is spread across the remaining territories. **Adjusted Sales** is the leftover after "
                "the first cap/floor pass. **# Terr in Band** counts territories that weren't capped "
                "or floored. **Adjustment/Terr** is the leftover divided across in-band territories. "
                "**Redistribution/Terr** comes from a second-pass redistribution of any overflow."
            )
            a1, a2, a3, a4 = st.columns(4)
            a1.metric("Adjusted Sales",
                      fmt(summary["R7_adjusted_sales"]),
                      delta="left after cap/floor", delta_color="off")
            a2.metric("# Terr in Band",
                      summary["R8_terr_in_band"],
                      delta="not capped/floored", delta_color="off")
            a3.metric("Adjustment / Terr",
                      fmt(summary["R9_adj_per_terr"]),
                      delta="added in-band", delta_color="off")
            a4.metric("Redistribution / Terr",
                      fmt(summary["T9_redist_per_terr"]),
                      delta="from overflow", delta_color="off")
            style_metric_cards(**LIGHT_CARDS)

            st.divider()

            # ── Stage-by-stage metric summary ─────────────────────
            st.subheader("Stage-by-Stage Metric Summary")
            st.markdown(
                "Each row represents one stage in the goal-setting process. As you read down the table, "
                "you can see how the territory goals evolve — from raw historical sales, through the "
                "weighted blend, volume smoothing, and finally the capped & redistributed result. "
                "**Max / Min / Average** show the spread across territories; **Total** shows the sum."
            )

            stat_cols = [
                ("MR3",                "Recent Period Sales",
                 "Sales from the most recent months — the primary baseline."),
                ("P3",                 "Older Period Sales",
                 "Sales from the months just before the recent period."),
                ("Current_Growth",     "Recent vs Older Growth %",
                 "How much sales grew (or shrank) recent vs older period."),
                ("Initial_Goal",       "Stage 1: Weighted Goal",
                 "First-pass goal using the recent/older weighting."),
                ("Initial_Growth",     "Stage 1: Implied Growth %",
                 "Growth implied by the Stage 1 goal vs recent sales."),
                ("Vol_Bal_Adj",        "Stage 2: Smoothing Adjustment",
                 "Adjustment from volume smoothing (positive = goal raised, negative = lowered)."),
                ("Goal_After_VolAdj",  "Stage 2: Smoothed Goal",
                 "Goal after applying the volume smoothing adjustment."),
                ("VolAdj_Growth",      "Stage 2: Implied Growth %",
                 "Growth implied by the smoothed goal."),
                ("Capped_Floored_Goal","Stage 3: After Cap & Floor",
                 "Goal after enforcing the max/min growth limits."),
                ("Capped_Growth",      "Stage 3: Implied Growth %",
                 "Growth after cap/floor enforcement."),
                ("Redistributed_Goal", "Stage 4: After Redistribution",
                 "Goal after redistributing excess/shortfall across territories."),
                ("Redist_Growth",      "Stage 4: Implied Growth %",
                 "Growth after redistribution."),
                ("Final_Goal",         "Final Territory Goal",
                 "The final goal assigned to each territory."),
                ("Final_Growth",       "Final Growth Target %",
                 "Final growth target each territory must achieve."),
            ]
            pct_cols_set = {"Current_Growth", "Initial_Growth", "VolAdj_Growth",
                            "Capped_Growth", "Redist_Growth", "Final_Growth"}

            stat_rows = []
            for key, label, desc in stat_cols:
                s = result[key].dropna()
                if s.empty:
                    stat_rows.append({"Stage / Metric": label, "What it shows": desc,
                                      "Max": "—", "Min": "—", "Average": "—", "Total": "—"})
                    continue
                if key in pct_cols_set:
                    stat_rows.append({
                        "Stage / Metric": label,
                        "What it shows":  desc,
                        "Max":            f"{s.max():+.2%}",
                        "Min":            f"{s.min():+.2%}",
                        "Average":        f"{s.mean():+.2%}",
                        "Total":          "—",
                    })
                else:
                    stat_rows.append({
                        "Stage / Metric": label,
                        "What it shows":  desc,
                        "Max":            f"{s.max():,.1f}",
                        "Min":            f"{s.min():,.1f}",
                        "Average":        f"{s.mean():,.1f}",
                        "Total":          f"{s.sum():,.1f}",
                    })
            stats_df = pd.DataFrame(stat_rows)
            st.dataframe(stats_df, use_container_width=True, hide_index=True)

            st.divider()

            # ── Territory Detail ──────────────────────────────────
            st.subheader("Per-Territory Breakdown")
            st.markdown(
                "Every territory's complete goal-setting journey, side by side. "
                "Read left → right to see how each territory's goal is built through the four stages. "
                "The **Final Territory Goal** column is what each rep will be measured against."
            )

            detail = result[[
                "Territory ID", "Territory Name", "Flat_Goal",
                "MR3", "P3", "Current_Growth",
                "Initial_Goal", "Initial_Growth",
                "Vol_Bal_Adj", "Goal_After_VolAdj", "VolAdj_Growth",
                "Capped_Floored_Goal", "Capped_Growth",
                "Redistributed_Goal", "Redist_Growth",
                "Final_Goal", "Final_Growth",
            ]].copy()

            mr_col_name = f"Recent {params['mr3_window']}-Mo"
            p_col_name  = f"Older {params['p3_window']}-Mo"

            detail.columns = [
                "Territory ID", "Territory Name", "Equal Share",
                mr_col_name, p_col_name, "Recent vs Older %",
                "Stage 1: Weighted Goal", "Stage 1 Growth %",
                "Smoothing Adj", "Stage 2: Smoothed Goal", "Stage 2 Growth %",
                "Stage 3: After Cap/Floor", "Stage 3 Growth %",
                "Stage 4: After Redistrib", "Stage 4 Growth %",
                "Final Territory Goal", "Final Growth %",
            ]

            num_cols = ["Equal Share", mr_col_name, p_col_name,
                        "Stage 1: Weighted Goal", "Smoothing Adj", "Stage 2: Smoothed Goal",
                        "Stage 3: After Cap/Floor", "Stage 4: After Redistrib", "Final Territory Goal"]
            pct_labels = ["Recent vs Older %", "Stage 1 Growth %", "Stage 2 Growth %",
                          "Stage 3 Growth %", "Stage 4 Growth %", "Final Growth %"]

            display = detail.copy()
            for c in num_cols:
                display[c] = display[c].map(lambda x: f"{x:,.1f}" if pd.notna(x) else "—")
            for c in pct_labels:
                display[c] = display[c].apply(lambda x: f"{x:+.2%}" if pd.notna(x) else "—")

            totals_row = {
                "Territory ID":               "TOTAL",
                "Territory Name":             "",
                "Equal Share":                f"{detail['Equal Share'].sum():,.1f}",
                mr_col_name:                  f"{detail[mr_col_name].sum():,.1f}",
                p_col_name:                   f"{detail[p_col_name].sum():,.1f}",
                "Recent vs Older %":          "—",
                "Stage 1: Weighted Goal":     f"{detail['Stage 1: Weighted Goal'].sum():,.1f}",
                "Stage 1 Growth %":           "—",
                "Smoothing Adj":              f"{detail['Smoothing Adj'].sum():,.1f}",
                "Stage 2: Smoothed Goal":     f"{detail['Stage 2: Smoothed Goal'].sum():,.1f}",
                "Stage 2 Growth %":           "—",
                "Stage 3: After Cap/Floor":   f"{detail['Stage 3: After Cap/Floor'].sum():,.1f}",
                "Stage 3 Growth %":           "—",
                "Stage 4: After Redistrib":   f"{detail['Stage 4: After Redistrib'].sum():,.1f}",
                "Stage 4 Growth %":           "—",
                "Final Territory Goal":       f"{detail['Final Territory Goal'].sum():,.1f}",
                "Final Growth %":             "—",
            }
            display = pd.concat([display, pd.DataFrame([totals_row])], ignore_index=True)
            st.dataframe(display, use_container_width=True, hide_index=True)

            st.divider()

            # ── Visual 1: Final Growth Distribution (sorted bar) ──
            st.subheader("Growth Target Distribution Across Territories")
            st.markdown(
                "Each bar is one territory, sorted from highest to lowest growth target. "
                "Use this to spot which territories are being asked to grow most aggressively "
                "and which are getting flat or declining goals. The dashed lines mark your "
                "Maximum / Minimum growth limits."
            )

            growth_df = result.dropna(subset=["Final_Growth"]).copy()
            growth_df["Final_Growth_pct"] = growth_df["Final_Growth"] * 100
            growth_df = growth_df.sort_values("Final_Growth_pct", ascending=False)

            axis_color  = "#f1f5f9" if is_dark else "#0f172a"
            grid_color  = "#1f2937" if is_dark else "#e2e8f0"
            tooltip_bg  = "#1e293b" if is_dark else "#ffffff"
            tooltip_fg  = "#f1f5f9" if is_dark else "#0f172a"
            tooltip_brd = "#334155" if is_dark else "#cbd5e1"

            fallback_on = summary.get("fallback_triggered", False)
            user_cap    = summary.get("user_cap",   growth_cap)
            user_floor  = summary.get("user_floor", growth_floor)

            # Color bars by zone:
            #   blue   = within user-set cap/floor (normal allocation)
            #   orange = pushed above user-set cap (fallback uplift)
            #   red    = pushed below user-set floor (fallback cutback)
            bar_colors = []
            for g in growth_df["Final_Growth_pct"]:
                if g > user_cap * 100 + 0.01:
                    bar_colors.append("#f97316")   # exceeded cap
                elif g < user_floor * 100 - 0.01:
                    bar_colors.append("#ef4444")   # below floor
                elif abs(g - user_cap * 100) < 0.01:
                    bar_colors.append("#f97316")   # exactly at cap
                elif abs(g - user_floor * 100) < 0.01:
                    bar_colors.append("#ef4444")   # exactly at floor
                else:
                    bar_colors.append("#2563eb")   # within band

            fig_growth = go.Figure(go.Bar(
                x=growth_df["Territory Name"],
                y=growth_df["Final_Growth_pct"],
                marker_color=bar_colors,
                marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Growth Target: %{y:.2f}%<extra></extra>",
            ))
            # User-set cap/floor reference lines (always shown)
            fig_growth.add_hline(y=user_cap * 100, line_dash="dash", line_color="#f97316",
                                 line_width=1.5,
                                 annotation_text=f"Max {user_cap:.0%}",
                                 annotation_position="top right",
                                 annotation_font_color=axis_color)
            fig_growth.add_hline(y=user_floor * 100, line_dash="dash", line_color="#ef4444",
                                 line_width=1.5,
                                 annotation_text=f"Min {user_floor:.0%}",
                                 annotation_position="bottom right",
                                 annotation_font_color=axis_color)
            fig_growth.update_layout(
                template="plotly_dark" if is_dark else "plotly_white",
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                height=460,
                margin=dict(l=70, r=20, t=30, b=160),
                xaxis=dict(
                    title=dict(text="Territory",
                               font=dict(color=axis_color, size=14, family="Calibri"),
                               standoff=18),
                    tickangle=-45,
                    color=axis_color, tickfont=dict(color=axis_color, size=10),
                    gridcolor=grid_color, linecolor=grid_color, showgrid=False,
                ),
                yaxis=dict(
                    title=dict(text="Final Growth Target (%)",
                               font=dict(color=axis_color, size=14, family="Calibri"),
                               standoff=12),
                    ticksuffix="%",
                    color=axis_color, tickfont=dict(color=axis_color, size=11),
                    gridcolor=grid_color, linecolor=grid_color, zeroline=True,
                    zerolinecolor=grid_color,
                ),
                font=dict(color=axis_color),
                hoverlabel=dict(bgcolor=tooltip_bg, bordercolor=tooltip_brd,
                                font=dict(color=tooltip_fg, size=12)),
            )

            # Color legend (changes wording when fallback is on)
            lc1, lc2, lc3 = st.columns(3)
            lc1.markdown("🟦 **Within growth band** — normal allocation")
            if fallback_on:
                lc2.markdown("🟧 **Pushed above max** — fallback uplift to match National Goal")
                lc3.markdown("🟥 **Pushed below min** — fallback cutback to match National Goal")
            else:
                lc2.markdown("🟧 **Capped at maximum** — limited by Max Growth setting")
                lc3.markdown("🟥 **Floored at minimum** — limited by Min Growth setting")
            st.plotly_chart(fig_growth, use_container_width=True)

            st.divider()

            # ── Visual 2: Territory Size vs Growth Target ─────────
            st.subheader("Territory Size vs Growth Target")
            st.markdown(
                "Helps you check **fairness**: are smaller territories being asked to grow at a similar "
                "rate as larger ones? Each bubble is a territory — its size shows the goal amount, "
                "the x-axis shows current sales, and the y-axis shows the growth % asked of it. "
                "A wide vertical spread at any sales level means growth expectations vary a lot for "
                "similarly-sized territories."
            )

            scatter_df = result.dropna(subset=["MR3", "Final_Growth"]).copy()
            scatter_df["Final_Growth_pct"] = scatter_df["Final_Growth"] * 100
            scatter_df["BubbleSize"] = (scatter_df["Final_Goal"] /
                                        scatter_df["Final_Goal"].max() * 40 + 8)

            # Color bubbles by zone — based on USER-set cap/floor so fallback
            # uplift/cutback is clearly visible
            bubble_colors = []
            for g in scatter_df["Final_Growth_pct"]:
                if g > user_cap * 100 + 0.01:
                    bubble_colors.append("#f97316")
                elif g < user_floor * 100 - 0.01:
                    bubble_colors.append("#ef4444")
                elif abs(g - user_cap * 100) < 0.01:
                    bubble_colors.append("#f97316")
                elif abs(g - user_floor * 100) < 0.01:
                    bubble_colors.append("#ef4444")
                else:
                    bubble_colors.append("#2563eb")

            fig_bubble = go.Figure(go.Scatter(
                x=scatter_df["MR3"],
                y=scatter_df["Final_Growth_pct"],
                mode="markers",
                marker=dict(size=scatter_df["BubbleSize"],
                            color=bubble_colors,
                            opacity=0.75,
                            line=dict(width=1.5, color="white" if is_dark else "#0f172a")),
                text=scatter_df["Territory Name"],
                customdata=np.stack([scatter_df["Final_Goal"],
                                      scatter_df["MR3"]], axis=-1),
                hovertemplate=(
                    "<b>%{text}</b><br>"
                    "Recent Sales: %{x:,.0f}<br>"
                    "Growth Target: %{y:.2f}%<br>"
                    "Final Goal: %{customdata[0]:,.0f}<extra></extra>"
                ),
            ))
            fig_bubble.add_hline(y=user_cap * 100, line_dash="dash", line_color="#f97316",
                                 line_width=1.2,
                                 annotation_text=f"Max {user_cap:.0%}",
                                 annotation_position="top right",
                                 annotation_font_color=axis_color)
            fig_bubble.add_hline(y=user_floor * 100, line_dash="dash", line_color="#ef4444",
                                 line_width=1.2,
                                 annotation_text=f"Min {user_floor:.0%}",
                                 annotation_position="bottom right",
                                 annotation_font_color=axis_color)
            fig_bubble.update_layout(
                template="plotly_dark" if is_dark else "plotly_white",
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                height=480,
                margin=dict(l=70, r=20, t=30, b=70),
                xaxis=dict(
                    title=dict(text="Recent Period Sales (Territory Size)",
                               font=dict(color=axis_color, size=14, family="Calibri"),
                               standoff=12),
                    tickformat=",.0f",
                    color=axis_color, tickfont=dict(color=axis_color, size=11),
                    gridcolor=grid_color, linecolor=grid_color,
                ),
                yaxis=dict(
                    title=dict(text="Growth Target (%)",
                               font=dict(color=axis_color, size=14, family="Calibri"),
                               standoff=12),
                    ticksuffix="%",
                    color=axis_color, tickfont=dict(color=axis_color, size=11),
                    gridcolor=grid_color, linecolor=grid_color,
                    zeroline=True, zerolinecolor=grid_color,
                ),
                font=dict(color=axis_color),
                hoverlabel=dict(bgcolor=tooltip_bg, bordercolor=tooltip_brd,
                                font=dict(color=tooltip_fg, size=12)),
            )
            st.plotly_chart(fig_bubble, use_container_width=True)
            # High-contrast hint (st.caption is too washed out in light mode)
            hint_color = "#cbd5e1" if is_dark else "#0f172a"
            st.markdown(
                f"<div style='font-size: 0.9rem; color: {hint_color}; "
                f"margin-top: 6px; font-weight: 500;'>"
                f"💡 Bubble size reflects the final goal amount — larger bubble = "
                f"larger absolute goal.</div>",
                unsafe_allow_html=True,
            )

            st.divider()

            # ── Navigation to Final Allocation ────────────────────
            st.info(
                "✅ Your weighted allocation is ready. Head to **Final Allocation** to review "
                "the finalised territory goals and download them, or continue tuning parameters above."
            )
            nav_l, nav_r = st.columns([1, 1])
            with nav_l:
                if st.button("Next: Back Testing →", type="primary", use_container_width=True):
                    go_to_tab(4)

        # ── FAIR SHARE MODEL ──────────────────────────────────────
        elif selected == 1:
            st.subheader("Fair Share Allocation")
            st.markdown(
                "Distributes the National Goal across territories based on each territory's "
                "**share of recent sales**. Larger-selling territories get larger goals, in proportion."
            )
            mr3_win = 3

            fs_df, total_mr3, ng_growth, mr3_months = compute_fair_share(df, nation_goal, mr3_win)

            # Persist results
            st.session_state.fair_share_result_df  = fs_df
            st.session_state.selected_model_label  = "Fair Share Model"

            s1, s2, s3, s4 = st.columns(4)
            s1.metric("National Goal",      fmt(nation_goal))
            s2.metric(f"Recent {mr3_win}-Mo Sales", fmt(total_mr3))
            s3.metric("Implied Growth",     pct_fmt(ng_growth),
                      delta=f"vs recent {mr3_win}-mo baseline",
                      delta_color="normal" if (ng_growth or 0) >= 0 else "inverse")
            s4.metric("Territories",        len(fs_df))
            style_metric_cards(**DARK_CARDS)

            disp = fs_df[["Territory ID", "Territory Name",
                           "MR3", "Share_%", "Goal", "Goal_Growth_%"]].copy()
            disp.columns = ["Territory ID", "Territory Name",
                            f"Recent {mr3_win}-Mo Sales", "Share %", "Territory Goal", "Growth Target %"]
            disp[f"Recent {mr3_win}-Mo Sales"] = disp[f"Recent {mr3_win}-Mo Sales"].map("{:,.1f}".format)
            disp["Share %"]            = disp["Share %"].map("{:.2%}".format)
            disp["Territory Goal"]     = disp["Territory Goal"].map("{:,.1f}".format)
            disp["Growth Target %"]    = disp["Growth Target %"].apply(
                lambda x: f"{x:+.2%}" if pd.notna(x) else "—"
            )
            totals = pd.DataFrame([{
                "Territory ID":       "TOTAL",
                "Territory Name":     "",
                f"Recent {mr3_win}-Mo Sales": f"{total_mr3:,.1f}",
                "Share %":            "100.00%",
                "Territory Goal":     f"{nation_goal:,.1f}",
                "Growth Target %":    pct_fmt(ng_growth),
            }])

            st.subheader("Per-Territory Breakdown")
            st.markdown(
                "Each territory's share of recent sales determines what % of the National Goal "
                "it receives. The growth target shows how much more (or less) each territory "
                "needs to deliver compared to its recent baseline."
            )
            st.dataframe(pd.concat([disp, totals], ignore_index=True),
                         use_container_width=True, hide_index=True)

            st.divider()
            st.download_button(
                "⬇️ Download Fair Share Allocation",
                data=to_excel(fs_df.rename(columns={
                    "MR3":           f"Recent {mr3_win}-Mo Sales",
                    "Share_%":       "Share %",
                    "Goal":          "Final Territory Goal",
                    "Goal_Growth_%": "Final Growth %",
                }), "Fair_Share_Allocation",
                    percent_cols=["Share %", "Final Growth %"],
                    currency_cols=[f"Recent {mr3_win}-Mo Sales", "Final Territory Goal"],
                    title=f"Fair Share Allocation — National Goal {fmt(nation_goal)}"),
                file_name="fair_share_allocation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── EQUAL ALLOCATION ──────────────────────────────────────
        elif selected == 2:
            st.subheader("Equal Allocation")
            st.markdown(
                "Splits a chosen portion of the National Goal **equally across all territories**, "
                "regardless of size or performance. Useful for new product launches where historical "
                "data isn't predictive."
            )

            eq_col, _ = st.columns([1, 2])
            with eq_col:
                eq_pct = st.number_input(
                    "Portion to Allocate Equally (%)",
                    min_value=0.0, max_value=100.0,
                    value=float(st.session_state.eq_pct), step=0.5, format="%.1f",
                    help="What % of the National Goal to split equally across territories. "
                         "100% means every territory gets exactly the same goal.",
                )
                st.session_state.eq_pct = eq_pct

            equal_pool  = nation_goal * eq_pct / 100
            per_terr_eq = equal_pool / n_terr if n_terr > 0 else 0

            e1, e2, e3 = st.columns(3)
            e1.metric("Total Pool to Distribute", fmt(equal_pool))
            e2.metric("Goal per Territory",       fmt(per_terr_eq))
            e3.metric("Portion Allocated",        f"{eq_pct}%")
            style_metric_cards(**DARK_CARDS)

            eq_df = compute_equal_allocation(df, nation_goal, eq_pct)

            # Persist results
            st.session_state.equal_result_df       = eq_df
            st.session_state.selected_model_label  = "Equal Allocation Model"

            disp2 = eq_df.copy()
            disp2 = disp2[["Territory ID", "Territory Name", "Final_Goal", "Share_%"]]
            disp2["Final_Goal"] = disp2["Final_Goal"].map("{:,.1f}".format)
            disp2["Share_%"]    = disp2["Share_%"].map("{:.2%}".format)
            disp2.columns = ["Territory ID", "Territory Name", "Territory Goal", "Share %"]
            totals_eq = pd.DataFrame([{
                "Territory ID":   "TOTAL",
                "Territory Name": "",
                "Territory Goal": f"{eq_df['Final_Goal'].sum():,.1f}",
                "Share %":        "100.00%",
            }])

            st.subheader("Per-Territory Breakdown")
            st.markdown(
                "Since the allocation is equal, every territory receives the same goal amount."
            )
            st.dataframe(pd.concat([disp2, totals_eq], ignore_index=True),
                         use_container_width=True, hide_index=True)

            st.divider()
            # Reorder columns so "Share %" is last in the Excel file too
            eq_df_export = eq_df[["Territory ID", "Territory Name", "Final_Goal", "Share_%"]].rename(
                columns={
                    "Final_Goal": "Final Territory Goal",
                    "Share_%":    "Share %",
                }
            )
            st.download_button(
                "⬇️ Download Equal Allocation",
                data=to_excel(eq_df_export, "Equal_Allocation",
                    percent_cols=["Share %"],
                    currency_cols=["Final Territory Goal"],
                    title=f"Equal Allocation — National Goal {fmt(nation_goal)} ({eq_pct}% equal share)"),
                file_name="equal_allocation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ═══════════════════════════════════════════════════════════════════
# TAB 3 — FINAL ALLOCATION (Weighted Model only)
# ═══════════════════════════════════════════════════════════════════
elif active == 3:
    st.subheader("📊 Final Allocation")
    st.markdown(
        "Review and download the finalised territory goals from the **Weighted Model**. "
        "Once you're satisfied, proceed to **Back Testing** to validate the model against past data."
    )

    gs_product  = st.session_state.gs_product
    nation_goal = st.session_state.nation_goal
    params      = st.session_state.chosen_params or st.session_state.weighted_params

    if st.session_state.data is None or gs_product is None or nation_goal is None or params is None:
        st.warning(
            "⚠️ Run the Weighted Model, then Back Testing, before viewing Final Allocation "
            "(National Goal Setting → Weighted Model → Back Testing)."
        )
        if st.button("← Back to National Goal Setting"):
            go_to_tab(2)
        st.stop()

    # Re-run the final allocation on the SELECTED PRODUCT with the chosen parameters
    df_prod = st.session_state.data[st.session_state.data["Product"] == gs_product].copy()
    wr, ws  = compute_weighted_model(df_prod, nation_goal, params)
    wp      = params
    src     = st.session_state.chosen_params_src or "model"
    st.caption(f"🧪 Product: **{gs_product}**  ·  parameters from: **{src}**")
    
    # ── Headline KPIs ────────────────────────────────────────────
    st.subheader("Headline")
    st.markdown(
        "Top-line numbers showing the National Goal, the total amount allocated by the model, "
        "the historical baseline used, and the reconciliation status."
    )
    h1, h2, h3, h4 = st.columns(4)
    h1.metric("National Goal",        fmt(nation_goal))
    h2.metric("Total Allocated",      fmt(ws["final_total_goal"]))
    h3.metric(f"Recent {wp['mr3_window']}-Mo Baseline", fmt(ws["total_mr3"]))
    diff = ws["final_minus_nation"]
    h4.metric("Reconciliation Gap",
              f"{diff:+,.1f}",
              delta="match" if abs(diff) < 1 else "check params",
              delta_color="normal" if abs(diff) < 1 else "inverse")
    style_metric_cards(**DARK_CARDS)

    st.divider()

    # ── Goal distribution chart ──────────────────────────────────
    st.subheader("Goal Distribution")
    st.markdown(
        "How the total goal pool is split across territories, sorted from largest to smallest goal. "
        "The dotted line marks the average territory goal — bars above it carry above-average targets, "
        "bars below it carry less."
    )

    goal_df = wr[["Territory Name", "Final_Goal"]].copy()
    goal_df = goal_df.sort_values("Final_Goal", ascending=False)
    avg_goal = goal_df["Final_Goal"].mean()

    axis_color  = "#f1f5f9" if is_dark else "#0f172a"
    grid_color  = "#1f2937" if is_dark else "#e2e8f0"
    tooltip_bg  = "#1e293b" if is_dark else "#ffffff"
    tooltip_fg  = "#f1f5f9" if is_dark else "#0f172a"
    tooltip_brd = "#334155" if is_dark else "#cbd5e1"

    fig_alloc = go.Figure()
    fig_alloc.add_trace(go.Bar(
        x=goal_df["Territory Name"],
        y=goal_df["Final_Goal"],
        marker_color="#2563eb",
        name="Final Territory Goal",
        hovertemplate="<b>%{x}</b><br>Final Goal: %{y:,.0f}<extra></extra>",
    ))
    fig_alloc.add_hline(y=avg_goal, line_dash="dot", line_color="#f97316",
                        line_width=2,
                        annotation_text=f"Avg {avg_goal:,.0f}",
                        annotation_position="top right",
                        annotation_font_color=axis_color)
    fig_alloc.update_layout(
        template="plotly_dark" if is_dark else "plotly_white",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        height=460,
        margin=dict(l=80, r=20, t=30, b=160),
        xaxis=dict(
            title=dict(text="Territory",
                       font=dict(color=axis_color, size=14, family="Calibri"),
                       standoff=18),
            tickangle=-45, color=axis_color,
            tickfont=dict(color=axis_color, size=10),
            linecolor=grid_color, showgrid=False,
        ),
        yaxis=dict(
            title=dict(text="Final Territory Goal ($)",
                       font=dict(color=axis_color, size=14, family="Calibri"),
                       standoff=12),
            tickformat=",.0f",
            color=axis_color, tickfont=dict(color=axis_color, size=11),
            gridcolor=grid_color, linecolor=grid_color,
        ),
        font=dict(color=axis_color),
        hoverlabel=dict(bgcolor=tooltip_bg, bordercolor=tooltip_brd,
                        font=dict(color=tooltip_fg, size=12)),
    )
    st.plotly_chart(fig_alloc, use_container_width=True)

    st.divider()

    # ── Final territory goals table (simplified) ─────────────────
    st.subheader("Final Territory Goals")
    st.markdown(
        "The official goal sheet — what each territory will be measured against. "
        "**Final Territory Goal** is the target each rep needs to achieve, and "
        "**Final Growth %** shows how much more (or less) that is compared to their recent baseline."
    )

    final_tbl = wr[["Territory ID", "Territory Name", "Final_Goal", "Final_Growth"]].copy()
    final_tbl.columns = ["Territory ID", "Territory Name",
                          "Final Territory Goal", "Final Growth %"]
    final_tbl_disp = final_tbl.copy()
    final_tbl_disp["Final Territory Goal"] = final_tbl_disp["Final Territory Goal"].map("{:,.1f}".format)
    final_tbl_disp["Final Growth %"] = final_tbl_disp["Final Growth %"].apply(
        lambda x: f"{x:+.2%}" if pd.notna(x) else "—"
    )

    totals_fa = pd.DataFrame([{
        "Territory ID":            "TOTAL",
        "Territory Name":          "",
        "Final Territory Goal":    f"{ws['final_total_goal']:,.1f}",
        "Final Growth %":          "—",
    }])
    st.dataframe(pd.concat([final_tbl_disp, totals_fa], ignore_index=True),
                 use_container_width=True, hide_index=True)

    # ── Download ─────────────────────────────────────────────────
    st.divider()
    st.download_button(
        "⬇️ Download Final Territory Goals",
        data=to_excel(final_tbl, "Final_Goals",
                      percent_cols=["Final Growth %"],
                      currency_cols=["Final Territory Goal"],
                      title=f"Final Territory Goals — National Goal {fmt(nation_goal)}"),
        file_name="final_territory_goals.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

   # ── Workflow: another product? ───────────────────────────────
    st.divider()
    st.markdown("##### Workflow")

    completed = list(st.session_state.gs_completed or [])
    if gs_product not in completed:
        completed.append(gs_product)
    all_products = st.session_state.selected_products or [gs_product]
    remaining = [p for p in all_products if p not in completed]

    if remaining:
        st.info(f"Finished **{gs_product}**. Set goals for another product?")
        w_l, w_r = st.columns(2)
        with w_l:
            if st.button("Yes", type="primary", use_container_width=True):
                st.session_state.gs_completed = completed
                st.session_state.gs_product = None      # forces the picker again
                for k in ("nation_goal", "weighted_result_df", "chosen_params",
                          "bt_result_df", "bt_metrics", "opt_results_df", "opt_best_params"):
                    st.session_state[k] = None
                st.session_state.nation_goal_submitted = False
                st.session_state.ngs_model_tab = None
                st.session_state.bt_intro_seen = False
                st.session_state.bt_mode = None
                st.session_state.chosen_params_src = None
                go_to_tab(2)
        with w_r:
            if st.button("No", use_container_width=True):
                st.session_state.gs_completed = completed
                go_to_tab(0)
    else:
        st.success("🎉 All selected products have been goal-set.")
        if st.button("← Back to Instructions", type="primary"):
            st.session_state.gs_completed = completed
            go_to_tab(0)

    st.divider()
    if st.button("← Back to Back Testing", use_container_width=False):
        go_to_tab(4)
            
# ═══════════════════════════════════════════════════════════════════
# TAB 4 — BACK TESTING
# ═══════════════════════════════════════════════════════════════════
elif active == 4:
    st.subheader("🔁 Back Testing")

    if st.session_state.data is None:
        st.warning("⚠️ Please upload and validate your data first.")
        if st.button("← Back to Input & Validation"):
            go_to_tab(1)
        st.stop()

    gs_product = st.session_state.gs_product
    if gs_product is None:
        st.warning("⚠️ Pick a product on the National Goal Setting tab first.")
        if st.button("← Back to National Goal Setting"):
            go_to_tab(2)
        st.stop()

    # ── Introduction gate ──
    if not st.session_state.bt_intro_seen:
        st.markdown(
            "### Why back test?\n"
            "Before you trust the Weighted Model to set live goals, it helps to ask: "
            "**would it have set *better* goals last quarter than the ones we actually used?**\n\n"
            "Back testing answers that. It trains the model on history up to a past quarter, "
            "asks it to set goals for the quarter that followed, then scores those goals against "
            "what really happened — and compares them to the goals you originally set.\n\n"
            "**What it evaluates:** accuracy (how close goals were to actual sales), fairness "
            "(cap/floor behaviour), and stability (does the model hold up across several quarters, "
            "not just one lucky one).\n\n"
            "You can either tune the parameters yourself, or let the **Grid Search Optimizer** "
            "find the most stable settings automatically."
        )
        if st.button("Proceed to Back Testing →", type="primary"):
            st.session_state.bt_intro_seen = True
            st.rerun()
        st.stop()

    st.caption(f"🧪 Back testing product: **{gs_product}**")

    bt_df = st.session_state.data[st.session_state.data["Product"] == gs_product].copy()
    bt_df["Week"]    = pd.to_datetime(bt_df["Week"])
    bt_df["Quarter"] = bt_df["Week"].dt.to_period("Q")
    quarters_avail   = sorted(bt_df["Quarter"].dropna().unique())

    if len(quarters_avail) < 3:
        st.warning("⚠️ Need at least **3 quarters** of data to run a back test "
                   f"(currently have {len(quarters_avail)}). Upload more history or "
                   "widen the date filter on the Input & Validation tab.")
        st.stop()

    quarter_labels = [str(q) for q in quarters_avail]
    
    # ──────────────────────────────────────────────────────────────
    # Mode: manual entry vs grid-search optimizer
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### How would you like to back test?")
    bt_mode = st.radio("Mode",
                       ["Manual Parameter Entry", "Grid Search Optimizer"],
                       horizontal=True, key="bt_mode_radio")
    st.divider()

    if bt_mode == "Grid Search Optimizer":
        st.markdown("##### Grid Search Optimizer")
        st.markdown(
            "Searches parameter combinations across several historical quarters, scores each with "
            "a composite metric, and picks the most **stable** performer — not just the one that "
            "happens to win a single quarter."
        )

        quarters = sorted(bt_df["Week"].dt.to_period("Q").dropna().unique())
        max_q = st.number_input(
            "Quarters to back test across",
            min_value=1, max_value=max(1, len(quarters) - 1),
            value=min(3, len(quarters) - 1), step=1, key="opt_max_q",
            help="Rolling back tests across recent quarters. More = steadier but slower.")
        pairs = build_quarter_pairs(quarters, int(max_q))
        st.caption("Testing quarters: " + ", ".join(str(t) for t, _ in pairs))

        with st.expander("⚙️ Candidate ranges (coarse grid)", expanded=True):
            r1, r2 = st.columns(2)
            with r1:
                w_lo = st.number_input("MR Weight from %",  0.0, 100.0, 50.0, 5.0, key="opt_w_lo")
                w_hi = st.number_input("MR Weight to %",    0.0, 100.0, 90.0, 5.0, key="opt_w_hi")
                w_st = st.number_input("MR Weight step %",  1.0,  50.0, 10.0, 1.0, key="opt_w_st")
                v_lo = st.number_input("Vol Balance from %",0.0,  20.0,  0.0, 1.0, key="opt_v_lo")
                v_hi = st.number_input("Vol Balance to %",  0.0,  20.0, 10.0, 1.0, key="opt_v_hi")
                v_st = st.number_input("Vol Balance step %",1.0,  10.0,  2.0, 1.0, key="opt_v_st")
            with r2:
                f_lo = st.number_input("Floor from %", -50.0, 50.0, -20.0, 5.0, key="opt_f_lo")
                f_hi = st.number_input("Floor to %",   -50.0, 50.0,   0.0, 5.0, key="opt_f_hi")
                f_st = st.number_input("Floor step %",   1.0, 50.0,  10.0, 1.0, key="opt_f_st")
                c_lo = st.number_input("Cap from %",     0.0,100.0,  10.0, 5.0, key="opt_c_lo")
                c_hi = st.number_input("Cap to %",       0.0,100.0,  30.0, 5.0, key="opt_c_hi")
                c_st = st.number_input("Cap step %",     1.0, 50.0,  10.0, 1.0, key="opt_c_st")
            refine = st.checkbox("Refine around best (fine ±1% pass)", value=True, key="opt_refine")

        ranges = {
            "mr_weight":    _frange(w_lo, w_hi, w_st),
            "vol_balance":  _frange(v_lo, v_hi, v_st),
            "growth_floor": _frange(f_lo, f_hi, f_st),
            "growth_cap":   _frange(c_lo, c_hi, c_st),
        }
        n_combo = sum(1 for w in ranges["mr_weight"] for v in ranges["vol_balance"]
                      for f in ranges["growth_floor"] for c in ranges["growth_cap"] if f <= c)
        st.info(f"Coarse grid: **{n_combo} combinations** × **{len(pairs)} quarters** "
                f"= **{n_combo * len(pairs)} model runs**"
                + (" (+ a small fine pass)" if refine else "")
                + ". Large grids can take 20–60 seconds.")

        base_bt_params = {"hist_scale": 1.0, "flat_goal_pct": 0.0,
                          "mr3_window": 3, "p3_window": 3,
                          "mr_weight": 0.7, "p_weight": 0.3, "vol_balance": 0.03,
                          "growth_cap": 0.10, "growth_floor": -0.10}

        if st.button("▶️ Run Optimizer", type="primary", key="opt_run"):
            prog = st.progress(0.0)
            status = st.empty()
            def _cb(done, total):
                prog.progress(min(done / total, 1.0))
                status.caption(f"Evaluating combination {done} / {total} …")
            with st.spinner("Searching parameter space…"):
                res_df, best = grid_search(bt_df, pairs, base_bt_params, ranges,
                                           progress_cb=_cb, refine=refine)
            prog.empty(); status.empty()
            if res_df.empty or best is None:
                st.error("No valid combinations — widen the ranges or add more quarters.")
            else:
                st.session_state.opt_results_df  = res_df
                st.session_state.opt_best_params = best

        res_df = st.session_state.opt_results_df
        best   = st.session_state.opt_best_params
        if res_df is not None and best is not None:
            st.success("✅ Optimization complete — best (most stable) parameters below.")
            ob1, ob2, ob3, ob4 = st.columns(4)
            ob1.metric("MR Weight",    f"{best['mr_weight']*100:.1f}%")
            ob2.metric("Vol Balance",  f"{best['vol_balance']*100:.1f}%")
            ob3.metric("Growth Floor", f"{best['growth_floor']*100:.1f}%")
            ob4.metric("Growth Cap",   f"{best['growth_cap']*100:.1f}%")
            style_metric_cards(**DARK_CARDS)

            ren = {"mr_weight_%": "MR Wt %", "vol_balance_%": "Vol Bal %",
                   "growth_floor_%": "Floor %", "growth_cap_%": "Cap %",
                   "stability_score": "Stability", "avg_composite": "Avg Composite",
                   "avg_wmape": "Avg WMAPE", "std_wmape": "WMAPE Std",
                   "avg_within_tol": "Within ±10%", "avg_bias": "Bias",
                   "avg_cap_floor_hit": "Cap/Floor Hit"}
            show = res_df.head(15)[list(ren.keys())].rename(columns=ren).copy()
            for c in ["Stability", "Avg Composite", "Avg WMAPE", "WMAPE Std", "Bias"]:
                show[c] = show[c].map(lambda x: f"{x:.4f}")
            show["Within ±10%"]  = show["Within ±10%"].map(lambda x: f"{x:.0%}")
            show["Cap/Floor Hit"] = show["Cap/Floor Hit"].map(lambda x: f"{x:.0%}")
            st.markdown("**Top 15 parameter sets (lower Stability = better)**")
            st.dataframe(show, use_container_width=True, hide_index=True)

            st.download_button(
                "⬇️ Download Full Optimizer Results",
                data=to_excel(res_df, "Optimizer", title=f"Grid Search — {gs_product}"),
                file_name="optimizer_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            if st.button("Use best parameters → Final Allocation", type="primary", key="opt_use"):
                st.session_state.chosen_params     = dict(best)
                st.session_state.chosen_params_src = "optimizer"
                go_to_tab(3)
        st.stop()
    # ↓↓↓ Manual Parameter Entry continues below (unchanged structure) ↓↓↓

    # ──────────────────────────────────────────────────────────────
    # 1) Quarter selectors — Test Quarter (Recent) + Prior Quarter
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Choose the quarters")
    st.markdown(
        "**Test Quarter** is the most recent quarter we want to replay - its actual sales and "
        "originally-set goals are the answer key. **Prior Quarter** is the quarter immediately "
        "before that - its sales become the model's *P3* baseline. The Prior Quarter must "
        "come *before* the Test Quarter; we won't let you pick the same one for both."
    )

    bt_q1, bt_q2 = st.columns(2)
    with bt_q1:
        # Default: second-to-last quarter as Test
        default_test_idx = len(quarter_labels) - 1
        test_label = st.selectbox(
            "Test Quarter (recent - has the actual sales)",
            quarter_labels, index=default_test_idx, key="bt_test_q",
            help="The quarter being back-tested. Weekly sales/goals in this quarter are "
                 "summed up and treated as the actual outcome to score against.",
        )
    test_quarter = pd.Period(test_label, freq="Q")

    # Prior quarter must be strictly before the test quarter
    prior_options = [q for q in quarter_labels if pd.Period(q, freq="Q") < test_quarter]
    if not prior_options:
        st.error("There is no quarter before the chosen Test Quarter. Pick a later Test Quarter.")
        st.stop()

    with bt_q2:
        prior_label = st.selectbox(
            "Prior Quarter (the model's P3 baseline)",
            prior_options, index=len(prior_options) - 1, key="bt_prior_q",
            help="The quarter immediately before Test. Used as the P3 baseline. "
                 "Anything earlier is treated as further history.",
        )
    prior_quarter = pd.Period(prior_label, freq="Q")

    # Friendly explanation of what the model sees
    train_end = prior_quarter.end_time
    train_rows = int((bt_df["Week"] <= train_end).sum())
    test_rows  = int((bt_df["Quarter"] == test_quarter).sum())
    st.info(
        f"📚 **Training data:** everything on or before "
        f"**{train_end.strftime('%d %b %Y')}** ({train_rows:,} rows). The model will use "
        f"**{prior_quarter}** as its P3 baseline and earlier quarters as further context.  \n"
        f"🎯 **Test data (held out):** **{test_quarter}** "
        f"({test_rows:,} rows). The model never sees these sales while training."
    )

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # 2) Model parameters (mirror of Weighted Model)
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Model Parameters")
    st.markdown(
        "These are the same parameters as the Weighted Model on the National Goal Setting tab. "
        "Tune them to see how different settings would have performed on a *past* quarter - "
        "if a configuration works well in the back test, it's likely to work well going forward."
    )

    with st.expander("⚙️  Weighted Model parameters for the replay", expanded=True):

        # ── Row 1: Weights & scaling ──────────────────────────────
        st.markdown(
            "**MR Weight** controls how much the most-recent quarter influences the goal. "
            "**P Weight** is automatically `100% - MR Weight`. "
            "**Historical Sales Scale** scales the training sales up or down before any "
            "calculation begins (leave at 100% in normal use)."
        )
        bp1, bp2, bp3 = st.columns(3)
        with bp1:
            # MR Weight
            bt_mr_pct = st.number_input("MR Weight (Most Recent) %", 0.0, 100.0, 70.0, 0.5,
                                        format="%.1f", key="bt_mr_pct",
                                        help="Weight applied to the Most Recent quarter's sales.")
            bt_mr = bt_mr_pct / 100.0
        with bp2:
            bt_p = 1.0 - bt_mr
            st.metric("P Weight (Prior)", f"{bt_p * 100:.0f}%",
                      delta="auto = 100% − MR Weight", delta_color="off")
        with bp3:
            # Historical Sales Scale
            bt_hist_pct = st.number_input("Historical Sales Scale %", 0.0, 200.0, 100.0, 0.5,
                                          format="%.1f", key="bt_hist_pct",
                                          help="Scaling factor on the training sales.")
            bt_hist = bt_hist_pct / 100.0

        # ── Row 2: Smoothing + cap/floor ──────────────────────────
        st.markdown(
            "**Volume Balance** is a smoothing factor that gently pulls small territories up "
            "and large territories down toward the average. **Growth Cap** is the maximum "
            "allowed growth % per territory. **Growth Floor** is the minimum - use it to "
            "ensure no territory is given a declining goal."
        )
        bp4, bp5, bp6 = st.columns(3)
        with bp4:
            # Volume Balance
            bt_vol_pct = st.number_input("Volume Balance %", 0.0, 10.0, 3.0, 0.1,
                                         format="%.1f", key="bt_vol_pct",
                                         help="Higher = stronger smoothing toward the average.")
            bt_vol = bt_vol_pct / 100.0
        with bp5:
            # Growth Cap
            bt_cap_pct = st.number_input("Growth Cap (Max) %", 0.0, 100.0, 2.0, 0.5,
                                         format="%.1f", key="bt_cap_pct",
                                         help="Maximum allowed growth % per territory.")
            bt_cap = bt_cap_pct / 100.0
        with bp6:
            # Growth Floor
            bt_floor_pct = st.number_input("Growth Floor (Min) %", -50.0, 50.0, -2.0, 0.5,
                                           format="%.1f", key="bt_floor_pct",
                                           help="Minimum allowed growth % per territory.")
            bt_floor = bt_floor_pct / 100.0

        # ── Row 3: Flat goal + windows ────────────────────────────
        st.markdown(
            "**Flat Goal %** is the portion of the National Goal split equally across all "
            "territories *before* the weighted blend. **MR Window** and **P Window** are how "
            "many months count as Most Recent and Prior (default 3 each = one quarter)."
        )
        bp7, bp8, bp9 = st.columns(3)
        with bp7:
            # Flat Goal
            bt_flat_pct = st.number_input("Flat Goal %", 0.0, 100.0, 0.0, 0.5,
                                          format="%.1f", key="bt_flat_pct",
                                          help="Equal-share portion before weighting.")
            bt_flat = bt_flat_pct / 100.0
        with bp8:
            bt_mrw = st.number_input("MR Window (months)", 1, 12, 3, 1,
                                     key="bt_mrw",
                                     help="Most-recent months used as the MR baseline.")
        with bp9:
            bt_pw = st.number_input("P Window (months)", 1, 12, 3, 1,
                                    key="bt_pw",
                                    help="Prior months used as the P baseline.")

    bt_params = {
        "hist_scale":    bt_hist,
        "flat_goal_pct": bt_flat,
        "mr_weight":     bt_mr,
        "p_weight":      bt_p,
        "vol_balance":   bt_vol,
        "growth_cap":    bt_cap,
        "growth_floor":  bt_floor,
        "mr3_window":    int(bt_mrw),
        "p3_window":     int(bt_pw),
    }

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Run / persisted results
    # ──────────────────────────────────────────────────────────────
    if st.button("▶️ Run Back Test", type="primary", key="bt_run"):
        bt_result, bt_metrics, bt_sim_summary = compute_back_test(
            bt_df, test_quarter, prior_quarter, bt_params)
        if bt_result.empty or not bt_metrics:
            st.error("Back test could not run - the chosen quarters do not both contain data.")
            st.stop()
        st.session_state.bt_result_df    = bt_result
        st.session_state.bt_metrics      = bt_metrics
        st.session_state.bt_sim_summary  = bt_sim_summary
        st.session_state.bt_params_used  = bt_params

    # Use persisted results so theme-toggle / re-renders don't wipe them
    bt_result      = st.session_state.bt_result_df
    bt_metrics     = st.session_state.bt_metrics
    bt_sim_summary = st.session_state.bt_sim_summary

    if bt_result is None or bt_metrics is None:
        st.info("Set your quarters and parameters above, then click **Run Back Test**.")
        st.stop()

    winner = bt_metrics["winner"]

    # ──────────────────────────────────────────────────────────────
    # Verdict
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Verdict")
    if "Simulated" in winner:
        improvement = (bt_metrics["mape_original"] - bt_metrics["mape_simulated"])
        st.success(
            f"✅ **The Weighted Model wins.** Its simulated goals were closer to actual sales "
            f"than the goals originally set - error dropped by **{improvement:.1%}**. The model "
            f"would have been a better basis for goal-setting in {bt_metrics['test_quarter']}."
        )
    else:
        gap = (bt_metrics["mape_simulated"] - bt_metrics["mape_original"])
        st.warning(
            f"⚠️ **The originally-set goals were more accurate** for "
            f"{bt_metrics['test_quarter']} - the model's error was higher by **{gap:.1%}**. "
            f"Try tuning MR Weight, Cap/Floor, or pick a different quarter pair."
        )

    v1, v2, v3, v4 = st.columns(4)
    v1.metric("Territories Tested", f"{bt_metrics['n_territories']:,}")
    v2.metric("Actual Sales",       fmt(bt_metrics["total_actual"]))
    v3.metric("Original Goals",     fmt(bt_metrics["total_original"]))
    v4.metric("Simulated Goals",    fmt(bt_metrics["total_simulated"]))
    style_metric_cards(**DARK_CARDS)

    st.divider()

    adj_sales      = bt_sim_summary["R7_adjusted_sales"]
    adj_per_terr   = bt_sim_summary["R9_adj_per_terr"]
    redist_per_terr= bt_sim_summary["T9_redist_per_terr"]
    fallback_on    = bt_sim_summary.get("fallback_triggered", False)

    show_adjustment = adj_sales > 1 and bt_sim_summary["R8_terr_in_band"] > 0

    if show_adjustment:
        st.markdown("##### Adjustment & Redistribution")
        st.markdown(
            "Some territories hit the Growth Cap or Floor during the replay, freeing up "
            "budget that the model spread across the remaining in-band territories so the "
            "totals still add up to the National Goal."
        )
        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Adjusted Sales",
                  fmt(adj_sales),
                  delta="left after cap/floor", delta_color="off")
        a2.metric("# Terr in Band",
                  bt_sim_summary["R8_terr_in_band"],
                  delta="not capped/floored", delta_color="off")
        a3.metric("Adjustment / Terr",
                  fmt(adj_per_terr),
                  delta="added in-band", delta_color="off")
        a4.metric("Redistribution / Terr",
                  fmt(redist_per_terr),
                  delta="from overflow", delta_color="off")
        style_metric_cards(**LIGHT_CARDS)
        st.divider()
    elif fallback_on:
        # The model rebalanced everything proportionally — explain it briefly
        st.info(
            "ℹ️ The cap/floor pass would have allocated more than the National Goal, so the "
            "model applied a **proportional rebalance** across all territories to bring the "
            "total back in line. This is the same fallback used on the National Goal Setting "
            "tab when the goal is tight relative to projected growth."
        )
        st.divider()

    # ──────────────────────────────────────────────────────────────
    # Accuracy Metrics
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Accuracy Metrics — Original vs Simulated")
    st.markdown(
        "**MAPE** = on average, how far off each goal was from actual sales, as a percentage. "
        "Lower is better. **RMSE** = same idea but in rupees, with bigger misses penalized "
        "more harshly. Lower is better. **R²** = how well the pattern of goals matches the "
        "pattern of sales (0 = no relationship, 1 = perfect). Higher is better."
    )

    def _fmt_pct(x):  return "—" if pd.isna(x) else f"{x:.1%}"
    def _fmt_num(x):  return "—" if pd.isna(x) else f"{x:,.0f}"
    def _fmt_r2(x):   return "—" if pd.isna(x) else f"{x:.3f}"

    metric_tbl = pd.DataFrame([
        {"Metric": "MAPE (lower is better)",
         "Original":  _fmt_pct(bt_metrics["mape_original"]),
         "Simulated": _fmt_pct(bt_metrics["mape_simulated"]),
         "Better":    "Simulated" if (bt_metrics["mape_simulated"] or 9e9) < (bt_metrics["mape_original"] or 9e9) else "Original"},
        {"Metric": "RMSE (lower is better)",
         "Original":  _fmt_num(bt_metrics["rmse_original"]),
         "Simulated": _fmt_num(bt_metrics["rmse_simulated"]),
         "Better":    "Simulated" if (bt_metrics["rmse_simulated"] or 9e9) < (bt_metrics["rmse_original"] or 9e9) else "Original"},
        {"Metric": "R² Correlation (higher is better)",
         "Original":  _fmt_r2(bt_metrics["r2_original"]),
         "Simulated": _fmt_r2(bt_metrics["r2_simulated"]),
         "Better":    "Simulated" if (bt_metrics["r2_simulated"] or -1) > (bt_metrics["r2_original"] or -1) else "Original"},
    ])
    st.dataframe(metric_tbl, use_container_width=True, hide_index=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("MAPE — Original",  _fmt_pct(bt_metrics["mape_original"]))
    m2.metric("MAPE — Simulated", _fmt_pct(bt_metrics["mape_simulated"]),
              delta="model wins" if "Simulated" in winner else "original wins",
              delta_color="normal" if "Simulated" in winner else "inverse")
    m3.metric("R² — Original",    _fmt_r2(bt_metrics["r2_original"]))
    m4.metric("R² — Simulated",   _fmt_r2(bt_metrics["r2_simulated"]))
    style_metric_cards(**LIGHT_CARDS)

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Stage-by-Stage stats (Max / Min / Average / Total) — workbook rows 3-6
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Stage-by-Stage Statistics")
    st.markdown(
        "How the model's numbers spread across territories at each stage of the calculation. "
        "**Max / Min / Average** show the range; **Total** shows the sum. Useful for spotting "
        "outliers and confirming that the totals add up to the National Goal."
    )

    # bt_result has Actual_Sales, Original_Goal, Simulated_Goal, etc.
    stat_rows = []
    stat_specs = [
        ("Actual_Sales",     "Actual Sales (test quarter)",      "currency"),
        ("Original_Goal",    "Original Goal",                    "currency"),
        ("Simulated_Goal",   "Simulated Goal (Weighted Model)",  "currency"),
        ("Original_Attainment",  "Original Attainment %",        "pct"),
        ("Simulated_Attainment", "Simulated Attainment %",       "pct"),
        ("Original_APE",     "Original Absolute % Error",        "pct"),
        ("Simulated_APE",    "Simulated Absolute % Error",       "pct"),
        ("Original_Payout",  "Original Payout (₹)",              "currency"),
        ("Simulated_Payout", "Simulated Payout (₹)",             "currency"),
    ]
    for key, label, kind in stat_specs:
        s = bt_result[key].dropna()
        if s.empty:
            stat_rows.append({"Metric": label, "Max": "—", "Min": "—",
                              "Average": "—", "Total": "—"})
            continue
        if kind == "pct":
            stat_rows.append({
                "Metric":  label,
                "Max":     f"{s.max():.1%}",
                "Min":     f"{s.min():.1%}",
                "Average": f"{s.mean():.1%}",
                "Total":   "—",
            })
        else:
            stat_rows.append({
                "Metric":  label,
                "Max":     f"{s.max():,.0f}",
                "Min":     f"{s.min():,.0f}",
                "Average": f"{s.mean():,.0f}",
                "Total":   f"{s.sum():,.0f}",
            })
    st.dataframe(pd.DataFrame(stat_rows), use_container_width=True, hide_index=True)

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Goals vs Actual Sales chart
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Goals vs Actual Sales")
    st.markdown(
        "Each territory's actual sales (blue bars) against the originally-set goal (grey line) "
        "and the model's simulated goal (orange dashed line). The closer a goal line hugs the "
        "tops of the bars, the more accurate it was."
    )

    bt_axis  = "#f1f5f9" if is_dark else "#0f172a"
    bt_grid  = "#1f2937" if is_dark else "#e2e8f0"
    bt_tt_bg = "#1e293b" if is_dark else "#ffffff"
    bt_tt_fg = "#f1f5f9" if is_dark else "#0f172a"
    bt_tt_br = "#334155" if is_dark else "#cbd5e1"

    chart_bt = bt_result.head(30)
    fig_bt = go.Figure()
    fig_bt.add_trace(go.Bar(
        x=chart_bt["Territory Name"], y=chart_bt["Actual_Sales"],
        name="Actual Sales", marker_color="#2563eb",
        hovertemplate="<b>%{x}</b><br>Actual: %{y:,.0f}<extra></extra>",
    ))
    fig_bt.add_trace(go.Scatter(
        x=chart_bt["Territory Name"], y=chart_bt["Original_Goal"],
        name="Original Goal", mode="lines+markers",
        line=dict(color="#94a3b8", width=2.5), marker=dict(size=6),
        hovertemplate="<b>%{x}</b><br>Original Goal: %{y:,.0f}<extra></extra>",
    ))
    fig_bt.add_trace(go.Scatter(
        x=chart_bt["Territory Name"], y=chart_bt["Simulated_Goal"],
        name="Simulated Goal", mode="lines+markers",
        line=dict(color="#f97316", width=2.5, dash="dash"),
        marker=dict(size=6, symbol="diamond"),
        hovertemplate="<b>%{x}</b><br>Simulated Goal: %{y:,.0f}<extra></extra>",
    ))
    fig_bt.update_layout(
        template="plotly_dark" if is_dark else "plotly_white",
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        height=460, margin=dict(l=80, r=20, t=30, b=160),
        legend=dict(orientation="h", y=1.06, x=1, xanchor="right",
                    font=dict(color=bt_axis)),
        xaxis=dict(
            title=dict(text="Territory",
                       font=dict(color=bt_axis, size=14, family="Calibri"),
                       standoff=18),
            tickangle=-45, color=bt_axis,
            tickfont=dict(color=bt_axis, size=10),
            linecolor=bt_grid, showgrid=False,
        ),
        yaxis=dict(
            title=dict(text="Amount ($)",
                       font=dict(color=bt_axis, size=14, family="Calibri"),
                       standoff=12),
            tickformat=",.0f", color=bt_axis,
            tickfont=dict(color=bt_axis, size=11),
            gridcolor=bt_grid, linecolor=bt_grid,
        ),
        font=dict(color=bt_axis),
        hoverlabel=dict(bgcolor=bt_tt_bg, bordercolor=bt_tt_br,
                        font=dict(color=bt_tt_fg, size=12)),
    )
    st.plotly_chart(fig_bt, use_container_width=True)

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Attainment Distribution 
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Attainment Distribution")
    st.markdown(
        "How many territories landed in each attainment band. A healthy goal-setting model "
        "produces a distribution centered near **100%** with not too many extreme over- or "
        "under-attainers. If most reps land below 70% or above 130%, the goals were too "
        "aggressive or too soft."
    )

    # Build buckets: 60-70%, 70-80%, ... 140-150%
    bucket_edges  = np.arange(0.6, 1.55, 0.1)
    bucket_labels = [f"{int(bucket_edges[i]*100)}–{int(bucket_edges[i+1]*100)}%"
                     for i in range(len(bucket_edges) - 1)]
    orig_counts = [int(((bt_result["Original_Attainment"]  >= bucket_edges[i]) &
                        (bt_result["Original_Attainment"]  <  bucket_edges[i+1])).sum())
                   for i in range(len(bucket_edges) - 1)]
    sim_counts  = [int(((bt_result["Simulated_Attainment"] >= bucket_edges[i]) &
                        (bt_result["Simulated_Attainment"] <  bucket_edges[i+1])).sum())
                   for i in range(len(bucket_edges) - 1)]

    fig_dist = go.Figure()
    fig_dist.add_trace(go.Bar(x=bucket_labels, y=orig_counts,
                              name="Original", marker_color="#94a3b8",
                              hovertemplate="<b>%{x}</b><br>Original: %{y} territories<extra></extra>"))
    fig_dist.add_trace(go.Bar(x=bucket_labels, y=sim_counts,
                              name="Simulated", marker_color="#f97316",
                              hovertemplate="<b>%{x}</b><br>Simulated: %{y} territories<extra></extra>"))
    fig_dist.update_layout(
        template="plotly_dark" if is_dark else "plotly_white",
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        barmode="group", height=380,
        margin=dict(l=70, r=20, t=30, b=70),
        legend=dict(orientation="h", y=1.06, x=1, xanchor="right",
                    font=dict(color=bt_axis)),
        xaxis=dict(
            title=dict(text="Attainment Band",
                       font=dict(color=bt_axis, size=14, family="Calibri"),
                       standoff=12),
            color=bt_axis, tickfont=dict(color=bt_axis, size=11),
            linecolor=bt_grid, gridcolor=bt_grid,
        ),
        yaxis=dict(
            title=dict(text="# Territories",
                       font=dict(color=bt_axis, size=14, family="Calibri"),
                       standoff=12),
            color=bt_axis, tickfont=dict(color=bt_axis, size=11),
            gridcolor=bt_grid, linecolor=bt_grid,
        ),
        font=dict(color=bt_axis),
        hoverlabel=dict(bgcolor=bt_tt_bg, bordercolor=bt_tt_br,
                        font=dict(color=bt_tt_fg, size=12)),
    )
    st.plotly_chart(fig_dist, use_container_width=True)

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Payout simulation
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Payout Simulation")
    st.markdown(
        "Simulated incentive payout assuming each territory earns **attainment × ₹10,000** "
        "(capped at 200%). Budget = $10,000 × number of territories. A well-calibrated goal "
        "lands total payout close to budget - over-payout means goals were too easy; "
        "under-payout means goals were too hard."
    )
    p1, p2, p3 = st.columns(3)
    p1.metric("Budgeted Payout", fmt(bt_metrics["budget_payout"]))
    p2.metric("Payout — Original", fmt(bt_metrics["payout_original"]),
              delta=f"{(bt_metrics['payout_original']/bt_metrics['budget_payout']-1):+.1%} vs budget"
              if bt_metrics["budget_payout"] else None, delta_color="off")
    p3.metric("Payout — Simulated", fmt(bt_metrics["payout_simulated"]),
              delta=f"{(bt_metrics['payout_simulated']/bt_metrics['budget_payout']-1):+.1%} vs budget"
              if bt_metrics["budget_payout"] else None, delta_color="off")
    style_metric_cards(**DARK_CARDS)

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Per-territory detail 
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### Per-Territory Detail")
    st.markdown(
        "One row per territory. Original vs Simulated goals, attainment, earnings, payout, "
        "and per-row error metrics - the same layout as the workbook's main table."
    )

    detail_bt = bt_result[[
        "Territory ID", "Territory Name", "Actual_Sales",
        "Original_Goal", "Original_Attainment", "Original_Earnings", "Original_Payout",
        "Simulated_Goal", "Simulated_Attainment", "Simulated_Earnings", "Simulated_Payout",
        "Original_APE", "Simulated_APE",
    ]].copy()
    detail_bt.columns = [
        "Territory ID", "Territory Name", "Actual Sales",
        "Original Goal", "Original Attain %", "Original Earnings", "Original Payout",
        "Simulated Goal", "Simulated Attain %", "Simulated Earnings", "Simulated Payout",
        "Original Error %", "Simulated Error %",
    ]
    disp_bt = detail_bt.copy()
    money_cols = ["Actual Sales", "Original Goal", "Simulated Goal",
                  "Original Payout", "Simulated Payout"]
    pct_cols   = ["Original Attain %", "Simulated Attain %",
                  "Original Earnings", "Simulated Earnings",
                  "Original Error %", "Simulated Error %"]
    for c in money_cols:
        disp_bt[c] = disp_bt[c].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
    for c in pct_cols:
        disp_bt[c] = disp_bt[c].apply(lambda x: f"{x:.1%}" if pd.notna(x) else "—")

    totals_bt = pd.DataFrame([{
        "Territory ID": "TOTAL", "Territory Name": "",
        "Actual Sales":      f"{detail_bt['Actual Sales'].sum():,.0f}",
        "Original Goal":     f"{detail_bt['Original Goal'].sum():,.0f}",
        "Original Attain %": "—", "Original Earnings": "—",
        "Original Payout":   f"{detail_bt['Original Payout'].sum():,.0f}",
        "Simulated Goal":    f"{detail_bt['Simulated Goal'].sum():,.0f}",
        "Simulated Attain %":"—", "Simulated Earnings":"—",
        "Simulated Payout":  f"{detail_bt['Simulated Payout'].sum():,.0f}",
        "Original Error %":  _fmt_pct(bt_metrics["mape_original"]),
        "Simulated Error %": _fmt_pct(bt_metrics["mape_simulated"]),
    }])
    st.dataframe(pd.concat([disp_bt, totals_bt], ignore_index=True),
                 use_container_width=True, hide_index=True)

    st.download_button(
        "⬇️ Download Back Test Results",
        data=to_excel(detail_bt, "Back_Test",
                      percent_cols=["Original Attain %", "Simulated Attain %",
                                    "Original Earnings", "Simulated Earnings",
                                    "Original Error %", "Simulated Error %"],
                      currency_cols=money_cols,
                      title=f"Back Test — Test {bt_metrics['test_quarter']} · "
                            f"Prior {bt_metrics['prior_quarter']}"),
        file_name="back_test_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()

    # ──────────────────────────────────────────────────────────────
    # Back Testing Summary
    # ──────────────────────────────────────────────────────────────
    st.markdown("##### 📋 Back Testing Summary")
    st.markdown(
        "A readout of everything above - what happened, how accurate the "
        "model was, and (if it lost) which specific parameters to tune to make it better."
    )

    # ── Accuracy score ────────────────────────────────────────────
    # Convert MAPE → an intuitive "Accuracy %": 100% − average error.
    # Floor at 0% so a wild error doesn't go negative on the user.
    acc_original  = max(0.0, 1 - (bt_metrics["mape_original"]  or 0))
    acc_simulated = max(0.0, 1 - (bt_metrics["mape_simulated"] or 0))
    acc_gap       = acc_simulated - acc_original
    model_wins    = "Simulated" in winner

    # ── Diagnostic signals from the result data (not generic advice) ──
    n_terr_total = len(bt_result)

    # Cap/floor utilisation
    user_cap   = bt_sim_summary.get("user_cap",   bt_params["growth_cap"])
    user_floor = bt_sim_summary.get("user_floor", bt_params["growth_floor"])
    n_at_cap   = int(((bt_result["Simulated_Growth"] >= user_cap   - 1e-6)).sum())
    n_at_floor = int(((bt_result["Simulated_Growth"] <= user_floor + 1e-6)).sum())
    pct_at_cap   = n_at_cap   / n_terr_total if n_terr_total else 0
    pct_at_floor = n_at_floor / n_terr_total if n_terr_total else 0

    # Direction bias: did the model systematically over- or under-shoot?
    diff_total = bt_metrics["total_simulated"] - bt_metrics["total_actual"]
    bias_pct   = diff_total / bt_metrics["total_actual"] if bt_metrics["total_actual"] else 0

    # Size-vs-error correlation: are larger territories systematically harder?
    if n_terr_total >= 3 and bt_result["Actual_Sales"].std() > 0:
        size_err_corr = float(bt_result[["Actual_Sales", "Simulated_APE"]]
                              .corr().iloc[0, 1])
    else:
        size_err_corr = 0.0

    # ── Plain-English headline card ───────────────────────────────
    if model_wins:
        verdict_bg     = "#dcfce7" if not is_dark else "#14532d"
        verdict_border = "#22c55e"
        verdict_icon   = "✅"
        verdict_head   = "The Weighted Model would have set more accurate goals."
        verdict_body   = (
            f"Across **{n_terr_total} territories** in **{bt_metrics['test_quarter']}**, "
            f"the model's goals were on average **{acc_simulated:.1%} accurate** — "
            f"better than the originally-set goals at **{acc_original:.1%}** "
            f"(an improvement of **{acc_gap:+.1%}**). "
            f"Total error dropped from **{bt_metrics['mape_original']:.1%}** to "
            f"**{bt_metrics['mape_simulated']:.1%}**."
        )
    else:
        verdict_bg     = "#fef3c7" if not is_dark else "#7c2d12"
        verdict_border = "#f97316"
        verdict_icon   = "⚠️"
        verdict_head   = "The originally-set goals were more accurate this time."
        verdict_body   = (
            f"Across **{n_terr_total} territories** in **{bt_metrics['test_quarter']}**, "
            f"the model achieved **{acc_simulated:.1%} accuracy** vs the originals at "
            f"**{acc_original:.1%}** "
            f"(model is **{abs(acc_gap):.1%}** behind). "
            f"The diagnostic checks below suggest specific parameter changes that "
            f"should close the gap."
        )

    verdict_text_color = "#0f172a" if not is_dark else "#f1f5f9"
    st.markdown(
        f"<div style='background:{verdict_bg};border-left:5px solid {verdict_border};"
        f"border-radius:10px;padding:18px 22px;margin:10px 0 18px 0;"
        f"color:{verdict_text_color};font-size:15px;line-height:1.65;'>"
        f"<div style='font-size:17px;font-weight:700;margin-bottom:8px;'>"
        f"{verdict_icon} {verdict_head}</div>"
        f"<div>{verdict_body}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Headline accuracy KPIs ────────────────────────────────────
    sa1, sa2, sa3, sa4 = st.columns(4)
    sa1.metric("Model Accuracy",        f"{acc_simulated:.1%}",
               delta=f"{acc_gap:+.1%} vs original",
               delta_color="normal" if model_wins else "inverse")
    sa2.metric("Original Accuracy",     f"{acc_original:.1%}")
    sa3.metric("Avg Error — Model",     f"{bt_metrics['mape_simulated']:.1%}",
               delta="lower is better", delta_color="off")
    sa4.metric("Avg Error — Original",  f"{bt_metrics['mape_original']:.1%}",
               delta="lower is better", delta_color="off")
    style_metric_cards(**DARK_CARDS)

    # ── Side-by-side comparison in plain language ─────────────────
    st.markdown("**At a glance**")
    side_l, side_r = st.columns(2)
    with side_l:
        st.markdown(
            f"""
**Original Goals**
- Total set: **{fmt(bt_metrics['total_original'])}**
- Average error: **{bt_metrics['mape_original']:.1%}** per territory
- Pattern match with actual sales (R²): **{bt_metrics['r2_original']:.2f}**
- Payout vs budget: **{(bt_metrics['payout_original']/bt_metrics['budget_payout']-1):+.1%}**
            """
        )
    with side_r:
        st.markdown(
            f"""
**Simulated Goals (Weighted Model)**
- Total set: **{fmt(bt_metrics['total_simulated'])}**
- Average error: **{bt_metrics['mape_simulated']:.1%}** per territory
- Pattern match with actual sales (R²): **{bt_metrics['r2_simulated']:.2f}**
- Payout vs budget: **{(bt_metrics['payout_simulated']/bt_metrics['budget_payout']-1):+.1%}**
            """
        )

    # Diagnostic Readout
    st.markdown("**Diagnostic Readout**")
    st.markdown(
        "Observations from this back test - what the numbers say about your data and the "
        "model's behaviour. Use these to decide whether to trust the model, try a different "
        "quarter pair, or accept that this period had factors the model couldn't see."
    )

    observations = []  # list of (tone, headline, body) tuples
    # tone: "good" | "neutral" | "warn"

    # ── Observation 1: Cap utilisation ────────────────────────────
    if pct_at_cap >= 0.40:
        observations.append((
            "warn",
            f"{n_at_cap} of {n_terr_total} territories ({pct_at_cap:.0%}) were pinned at the Growth Cap.",
            f"With the cap at **{user_cap:.0%}**, the model judged a large share of "
            f"territories to deserve higher growth than the cap allows. The cap is doing "
            f"its job - preventing unrealistic targets - but it also means the model can't "
            f"reflect what the data 'wants' for those territories. This is a **policy choice**, "
            f"not a bug. If the original goals beat the model here, it's likely because they "
            f"were allowed to set growth above {user_cap:.0%}."
        ))
    elif pct_at_cap >= 0.20:
        observations.append((
            "neutral",
            f"{n_at_cap} of {n_terr_total} territories ({pct_at_cap:.0%}) hit the Growth Cap.",
            f"A moderate share of territories were capped at **{user_cap:.0%}**. This is normal "
            f"and expected - the cap exists to keep goals achievable. Just be aware that for "
            f"these territories, the model is producing the safest goal, not the most "
            f"accurate one."
        ))

    # ── Observation 2: Floor utilisation ──────────────────────────
    if pct_at_floor >= 0.40:
        observations.append((
            "warn",
            f"{n_at_floor} of {n_terr_total} territories ({pct_at_floor:.0%}) were pinned at the Growth Floor.",
            f"With the floor at **{user_floor:.0%}**, many territories are being forced to "
            f"hold flat or grow when their underlying data suggests decline. This protects reps "
            f"from defeatist goals but reduces the model's freedom to forecast realistically. "
            f"If accuracy matters more than morale for this analysis, consider whether some "
            f"territories are *genuinely* shrinking and need that reflected."
        ))

    # ── Observation 3: Systematic bias ────────────────────────────
    if abs(bias_pct) >= 0.05:
        if bias_pct > 0:
            observations.append((
                "neutral",
                f"The model's total was {bias_pct:+.1%} above actual sales.",
                f"In hindsight, the test quarter ({bt_metrics['test_quarter']}) underperformed "
                f"what the model would have predicted from prior trends. This usually means **the "
                f"market softened** - something external happened (slower growth, competitor, "
                f"macro factor). The model didn't know about it because it can only see "
                f"historical data. No parameter change can fix unknowable future events."
            ))
        else:
            observations.append((
                "neutral",
                f"The model's total was {bias_pct:+.1%} below actual sales.",
                f"The test quarter ({bt_metrics['test_quarter']}) outperformed what the model "
                f"would have predicted - sales accelerated beyond what the prior trend suggested. "
                f"This is **good news for the business but a tough call for goal-setting**: "
                f"the model is being conservative because the historical pattern was less "
                f"aggressive than what actually happened."
            ))

    # ── Observation 4: Size-vs-error ──────────────────────────────
    if abs(size_err_corr) >= 0.40:
        if size_err_corr > 0:
            observations.append((
                "neutral",
                f"Larger territories had higher forecast error (correlation {size_err_corr:+.2f}).",
                "Bigger territories were harder for the model to predict - they tend to swing "
                "more in absolute rupees. This is common and not necessarily a problem; just be "
                "aware that the model's confidence is lower for your biggest territories."
            ))
        else:
            observations.append((
                "neutral",
                f"Smaller territories had higher forecast error (correlation {size_err_corr:+.2f}).",
                "Smaller territories were harder for the model to predict - their sales can be "
                "noisier in percentage terms. Consider whether the smallest territories warrant a "
                "different goal-setting approach (e.g. Equal Allocation for territories below a "
                "size threshold)."
            ))

    # ── Observation 5: Outlier territories ────────────────────────
    rmse_avg_ratio = (bt_metrics["rmse_simulated"] /
                      (bt_metrics["total_actual"] / max(n_terr_total, 1) or 1))
    if rmse_avg_ratio >= 0.5 and bt_metrics["mape_simulated"] < 0.25:
        observations.append((
            "neutral",
            "A handful of territories are far off - most are accurate.",
            "Average error is moderate but RMSE is high, which means **a few territories are "
            "wildly off and the rest are fine**. Scroll up to the Per-Territory Detail table, "
            "sort by Simulated Error %, and look at the worst 2-3 rows. They may be new "
            "launches, transitions, or data-quality issues worth investigating outside the model."
        ))

    # ── Observation 6: Model on par with originals ────────────────
    if abs(acc_gap) < 0.03:
        observations.append((
            "good",
            f"Model accuracy is within {abs(acc_gap):.1%} of the originally-set goals.",
            "Essentially a tie. The Weighted Model produces goals that are as accurate as the "
            "ones a human set with full context - which is a strong endorsement of using the "
            "model going forward, since it's faster, repeatable, and removes individual bias."
        ))

    # ── Observation 7: Clean win ──────────────────────────────────
    if model_wins and acc_gap >= 0.05:
        observations.append((
            "good",
            f"Model beat the originals by {acc_gap:.1%}.",
            "The Weighted Model produced more accurate goals than what was actually set for "
            f"{bt_metrics['test_quarter']}. The current parameters look well-calibrated for this "
            "type of period."
        ))

    # ── Fallback when nothing notable ─────────────────────────────
    if not observations:
        observations.append((
            "neutral",
            "Nothing unusual detected.",
            "The back test ran cleanly and the standard diagnostics didn't flag anything "
            "notable. Try a different quarter pair to see how the model performs across "
            "multiple periods before drawing firm conclusions."
        ))

    # ── Render observations ───────────────────────────────────────
    tone_color = {
        "good":    ("#22c55e", "#dcfce7" if not is_dark else "#14532d"),
        "neutral": ("#3b82f6", "#dbeafe" if not is_dark else "#1e3a5f"),
        "warn":    ("#f97316", "#fef3c7" if not is_dark else "#7c2d12"),
    }
    tone_icon = {"good": "✅", "neutral": "📊", "warn": "⚠️"}

    for tone, headline, body in observations:
        accent, bg = tone_color[tone]
        icon = tone_icon[tone]
        st.markdown(
            f"<div style='background:{bg};border-left:4px solid {accent};"
            f"border-radius:8px;padding:14px 18px;margin:8px 0;"
            f"color:{verdict_text_color};font-size:14.5px;line-height:1.6;'>"
            f"<div style='font-weight:600;margin-bottom:6px;'>{icon} {headline}</div>"
            f"<div>{body}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── How to act on this ────────────────────────────────────────
    st.markdown("**How to use this back test**")
    st.markdown(
        "- **If the model won** - the current Weighted Model parameters are working. "
        "Use them on the National Goal Setting tab with confidence.\n"
        "- **If the model lost narrowly** - try running the back test on **a different "
        "quarter pair**. A single quarter can be misleading; if the model is consistently "
        "close across multiple back tests, it's reliable.\n"
        "- **If the model lost by a lot** - the test quarter probably had factors the model "
        "couldn't see (market shift, launch, disruption). The original goals likely benefited "
        "from human judgment about those factors. The model is best treated as a *starting "
        "point* that humans then adjust, not a final answer.\n"
        "- **Don't 'tune to win'** - widening the Growth Cap or shifting weights just to "
        "improve back test accuracy will produce unrealistic goals in live use. The model's "
        "parameters exist to keep goals **fair and achievable**, not to fit history perfectly."
    )

    # ── Bottom-line takeaway ──────────────────────────────────────
    if model_wins and acc_gap >= 0.05:
        bottom_line = (
            f"**Use the Weighted Model with confidence.** It beat the originally-set goals "
            f"by **{acc_gap:.1%}** on accuracy in this back test."
        )
    elif model_wins:
        bottom_line = (
            f"**The Weighted Model is roughly on par with - slightly better than - the "
            f"originally-set goals.** The improvement is modest ({acc_gap:+.1%}), so test "
            f"another quarter before committing."
        )
    elif acc_gap >= -0.03:
        bottom_line = (
            f"**The model is close to the originals** (just {abs(acc_gap):.1%} behind). "
            f"Apply the fixes above and re-run - it should overtake easily."
        )
    else:
        bottom_line = (
            f"**Don't deploy the model with these settings yet.** It's {abs(acc_gap):.1%} "
            f"behind the originals. Apply the fixes above (highest priority first), re-run "
            f"this back test, and aim for Model Accuracy ≥ Original Accuracy before using "
            f"the model for live goal-setting."
        )
    st.info(f"**Bottom line:** {bottom_line}")
    
    st.divider()
    st.markdown("##### Carry these parameters into Final Allocation")
    st.markdown("When you're happy with the manual settings above, send them forward.")
    cf_l, cf_r = st.columns(2)
    with cf_l:
        if st.button("Use these parameters → Final Allocation", type="primary",
                     use_container_width=True, key="bt_use_manual"):
            st.session_state.chosen_params     = dict(bt_params)
            st.session_state.chosen_params_src = "manual"
            go_to_tab(3)
    with cf_r:
        if st.button("← Back to National Goal Setting", use_container_width=True, key="bt_back"):
            go_to_tab(2)


